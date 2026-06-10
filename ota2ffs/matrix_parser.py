from __future__ import annotations

from pathlib import Path
from typing import Iterable, TypeAlias

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .matrix_model import PatternMatrix
from .utils import cell_text, is_text, normalize_angle, to_float


BlockStart: TypeAlias = tuple[int, int]


def parse_matrices_from_workbook(
    excel_path: str | Path,
    selected_sheets: Iterable[str] | None = None,
) -> list[PatternMatrix]:
    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    try:
        sheet_names = list(selected_sheets) if selected_sheets is not None else list(workbook.sheetnames)
        matrices: list[PatternMatrix] = []
        for sheet_name in sheet_names:
            if sheet_name not in workbook.sheetnames:
                continue
            ws = workbook[sheet_name]
            sheet_matrices = parse_v2_sheet(ws)
            if not sheet_matrices:
                sheet_matrices = parse_v1_sheet(ws)
            matrices.extend(sheet_matrices)
        return matrices
    finally:
        workbook.close()


def parse_v1_sheet(ws: Worksheet) -> list[PatternMatrix]:
    matrices: list[PatternMatrix] = []
    for block_name in ("Theta", "Phi"):
        block_start = _find_v1_block(ws, block_name)
        if block_start is None:
            continue
        matrix = _parse_v1_block(ws, block_start, block_name)
        if _has_matrix_data(matrix):
            matrices.append(matrix)
    if len({matrix.block_name for matrix in matrices}) < 2:
        return []
    return matrices


def parse_v2_sheet(ws: Worksheet) -> list[PatternMatrix]:
    matrices: list[PatternMatrix] = []
    for row in range(1, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
            if not is_text(ws.cell(row=row, column=column).value, "Polarization"):
                continue
            block_name = cell_text(ws.cell(row=row, column=column + 1).value)
            if block_name.casefold() not in {"theta", "phi", "total"}:
                continue
            matrix = _parse_v2_table(ws, row, column, block_name)
            if _has_matrix_data(matrix):
                matrices.append(matrix)
    if len({matrix.block_name.casefold() for matrix in matrices} & {"theta", "phi"}) < 2:
        return []
    return matrices


def _find_v1_block(ws: Worksheet, block_name: str) -> BlockStart | None:
    for row in range(1, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
            if (
                is_text(ws.cell(row=row, column=column).value, block_name)
                and is_text(ws.cell(row=row, column=column + 1).value, "Phi Angle")
                and is_text(ws.cell(row=row + 1, column=column + 1).value, "Theta Angle")
            ):
                return row, column
    return None


def _parse_v1_block(ws: Worksheet, start: BlockStart, block_name: str) -> PatternMatrix:
    start_row, start_column = start
    row_angle_column = start_column + 1
    first_value_column = start_column + 2
    col_angles, value_columns = _read_col_angles(ws, start_row, first_value_column)
    row_angles, values = _read_matrix_rows(ws, start_row + 2, row_angle_column, value_columns)
    return PatternMatrix(
        sheet_name=ws.title,
        block_name=block_name,
        row_label="Phi Angle",
        col_label="Theta Angle",
        row_angles=row_angles,
        col_angles=col_angles,
        values=values,
    )


def _parse_v2_table(ws: Worksheet, start_row: int, start_column: int, block_name: str) -> PatternMatrix:
    header_row = start_row + 1
    row_angle_column = start_column
    first_value_column = start_column + 1
    col_angles, value_columns = _read_col_angles(ws, header_row, first_value_column)
    row_angles, values = _read_matrix_rows(ws, start_row + 2, row_angle_column, value_columns)
    return PatternMatrix(
        sheet_name=ws.title,
        block_name=block_name,
        row_label="Phi Angle",
        col_label="Theta Angle",
        row_angles=row_angles,
        col_angles=col_angles,
        values=values,
    )


def _read_col_angles(ws: Worksheet, row: int, first_column: int) -> tuple[list[float], list[int]]:
    angles: list[float] = []
    columns: list[int] = []
    for column in range(first_column, ws.max_column + 1):
        angle = to_float(ws.cell(row=row, column=column).value)
        if angle is None:
            if angles:
                break
            continue
        angles.append(normalize_angle(angle))
        columns.append(column)
    return angles, columns


def _read_matrix_rows(
    ws: Worksheet,
    first_row: int,
    row_angle_column: int,
    value_columns: list[int],
) -> tuple[list[float], list[list[float]]]:
    row_angles: list[float] = []
    values: list[list[float]] = []
    saw_data = False

    for row in range(first_row, ws.max_row + 1):
        row_angle = to_float(ws.cell(row=row, column=row_angle_column).value)
        if row_angle is None:
            if saw_data:
                break
            continue

        row_angles.append(normalize_angle(row_angle))
        saw_data = True
        values.append([_numeric_or_zero(ws.cell(row=row, column=column).value) for column in value_columns])

    return row_angles, values


def _numeric_or_zero(value) -> float:
    number = to_float(value)
    return float(number) if number is not None else 0.0


def _has_matrix_data(matrix: PatternMatrix) -> bool:
    return bool(matrix.row_angles and matrix.col_angles and matrix.values)
