from __future__ import annotations

from dataclasses import dataclass

from openpyxl.worksheet.worksheet import Worksheet

from .utils import FarFieldSource, cell_text, contains_text, extract_number, is_text, normalize_angle, sorted_unique, to_float


@dataclass(slots=True)
class _V2Table:
    polarization: str
    theta_angles: list[float]
    phi_angles: list[float]
    values: dict[tuple[float, float], float]
    frequency_mhz: float | None


def is_v2_sheet(ws: Worksheet) -> bool:
    polarizations = {table.polarization for table in _find_tables(ws)}
    return {"theta", "phi"}.issubset(polarizations)


def parse_sheet(ws: Worksheet) -> list[FarFieldSource]:
    tables = {table.polarization: table for table in _find_tables(ws)}
    theta_table = tables.get("theta")
    phi_table = tables.get("phi")
    total_table = tables.get("total")

    if theta_table is None or phi_table is None:
        raise ValueError("未找到 V2 所需的 Theta/Phi 表格")

    sources = [
        FarFieldSource(
            sheet_name=ws.title,
            theta_angles=sorted_unique([*theta_table.theta_angles, *phi_table.theta_angles]),
            phi_angles=sorted_unique([*theta_table.phi_angles, *phi_table.phi_angles]),
            e_theta_db=theta_table.values,
            e_phi_db=phi_table.values,
            version="V2",
            frequency_mhz=theta_table.frequency_mhz or phi_table.frequency_mhz,
        )
    ]

    if total_table is not None:
        sources.append(
            FarFieldSource(
                sheet_name=ws.title,
                suffix="_total",
                theta_angles=total_table.theta_angles,
                phi_angles=total_table.phi_angles,
                e_theta_db=total_table.values,
                e_phi_db={},
                version="V2",
                frequency_mhz=total_table.frequency_mhz,
            )
        )

    return sources


def _find_tables(ws: Worksheet) -> list[_V2Table]:
    tables: list[_V2Table] = []
    for row in range(1, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
            if not is_text(ws.cell(row=row, column=column).value, "Polarization"):
                continue
            polarization = cell_text(ws.cell(row=row, column=column + 1).value).casefold()
            if polarization not in {"theta", "phi", "total"}:
                continue
            try:
                tables.append(_parse_table(ws, row, column, polarization))
            except ValueError:
                continue
    return tables


def _parse_table(ws: Worksheet, start_row: int, start_column: int, polarization: str) -> _V2Table:
    theta_columns = _read_theta_columns(ws, start_row + 1, start_column + 1)
    if not theta_columns:
        raise ValueError(f"V2 {polarization} 表格未找到 Theta 角度")

    theta_angles = [theta for _, theta in theta_columns]
    phi_angles: list[float] = []
    values: dict[tuple[float, float], float] = {}
    saw_data = False

    for row in range(start_row + 2, ws.max_row + 1):
        if is_text(ws.cell(row=row, column=start_column).value, "Polarization"):
            break

        phi = to_float(ws.cell(row=row, column=start_column).value)
        if phi is None:
            if saw_data:
                break
            continue

        phi = normalize_angle(phi)
        phi_angles.append(phi)
        saw_data = True

        for column, theta in theta_columns:
            db_value = to_float(ws.cell(row=row, column=column).value)
            if db_value is not None:
                values[(phi, theta)] = db_value

    if not phi_angles:
        raise ValueError(f"V2 {polarization} 表格未找到 Phi 角度")

    return _V2Table(
        polarization=polarization,
        theta_angles=theta_angles,
        phi_angles=phi_angles,
        values=values,
        frequency_mhz=_read_frequency(ws, start_row, start_column),
    )


def _read_theta_columns(ws: Worksheet, row: int, first_column: int) -> list[tuple[int, float]]:
    theta_columns: list[tuple[int, float]] = []
    for column in range(first_column, ws.max_column + 1):
        theta = to_float(ws.cell(row=row, column=column).value)
        if theta is None:
            if theta_columns:
                break
            continue
        theta_columns.append((column, normalize_angle(theta)))
    return theta_columns


def _read_frequency(ws: Worksheet, row: int, start_column: int) -> float | None:
    for column in range(start_column, ws.max_column + 1):
        if contains_text(ws.cell(row=row, column=column).value, "Freq"):
            frequency = extract_number(ws.cell(row=row, column=column + 1).value)
            if frequency is not None:
                return frequency

    for column in range(ws.max_column, start_column - 1, -1):
        frequency = extract_number(ws.cell(row=row, column=column).value)
        if frequency is not None:
            return frequency
    return None
