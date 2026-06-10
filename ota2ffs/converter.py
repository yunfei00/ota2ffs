from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from . import parser_v1, parser_v2
from .ffs_writer import write_ffs
from .report_writer import write_conversion_log
from .utils import FarFieldSource, frequency_to_hz, sanitize_filename


@dataclass(slots=True)
class ConversionResult:
    generated_files: list[Path] = field(default_factory=list)
    failures: dict[str, str] = field(default_factory=dict)
    log_lines: list[str] = field(default_factory=list)
    log_path: Path | None = None

    @property
    def generated_count(self) -> int:
        return len(self.generated_files)


def convert_excel(
    excel_path: str | Path,
    output_dir: str | Path,
    sheet_names: Iterable[str] | None = None,
    frequency_value: str | float | int | None = None,
    frequency_unit: str = "MHz",
    log_dir: str | Path | None = None,
) -> ConversionResult:
    excel_path = Path(excel_path)
    output_dir = Path(output_dir)
    excel_output_dir = output_dir / sanitize_filename(excel_path.stem)
    excel_output_dir.mkdir(parents=True, exist_ok=True)
    fallback_frequency_hz = frequency_to_hz(frequency_value, frequency_unit)

    result = ConversionResult()
    result.log_lines.append(f"Excel 文件: {excel_path}")
    result.log_lines.append(f"输出目录: {excel_output_dir}")
    if fallback_frequency_hz is not None:
        result.log_lines.append(f"界面频率[Hz]: {fallback_frequency_hz:g}")

    workbook = load_workbook(excel_path, data_only=True)
    selected_sheets = list(sheet_names) if sheet_names is not None else list(workbook.sheetnames)

    for sheet_name in selected_sheets:
        if sheet_name not in workbook.sheetnames:
            message = "sheet 不存在"
            result.failures[sheet_name] = message
            result.log_lines.append(f"[失败] {sheet_name}: {message}")
            continue

        ws = workbook[sheet_name]
        try:
            if parser_v2.is_v2_sheet(ws):
                sources = parser_v2.parse_sheet(ws)
                version = "V2"
            elif parser_v1.is_v1_sheet(ws):
                sources = [parser_v1.parse_sheet(ws)]
                version = "V1"
            else:
                raise ValueError("无法识别为 V1 或 V2 格式")

            before_count = len(result.generated_files)
            for source in sources:
                source_frequency_hz = _frequency_hz_for_source(source, fallback_frequency_hz)
                result.generated_files.append(write_ffs(source, excel_output_dir, "RX", source_frequency_hz))
                result.generated_files.append(write_ffs(source, excel_output_dir, "TX", source_frequency_hz))

            produced = len(result.generated_files) - before_count
            result.log_lines.append(f"[成功] {sheet_name}: {version}, 生成 {produced} 个文件")
        except Exception as exc:
            message = str(exc)
            result.failures[sheet_name] = message
            result.log_lines.append(f"[失败] {sheet_name}: {message}")

    workbook.close()

    result.log_lines.append(f"生成文件总数: {result.generated_count}")
    result.log_lines.append(f"失败 sheet 数量: {len(result.failures)}")
    result.log_path = write_conversion_log(result.log_lines, log_dir)
    return result


def _frequency_hz_for_source(source: FarFieldSource, fallback_frequency_hz: float | None) -> float:
    if source.frequency_mhz is not None:
        return frequency_to_hz(source.frequency_mhz, "MHz") or 0.0
    if fallback_frequency_hz is None:
        raise ValueError("缺少频率信息")
    return fallback_frequency_hz
