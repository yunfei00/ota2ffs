from __future__ import annotations

import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Iterable
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.chart import RadarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .matrix_model import PatternMatrix
from .matrix_parser import parse_matrices_from_workbook


# Excel's default inserted chart size: 360 pt x 216 pt.
CHART_WIDTH = 12.7
CHART_HEIGHT = 7.62
EXCEL_NATIVE_RADAR_STYLE = 2
SERIES_LINE_WIDTH_EMU = "28575"
GRID_LINE_WIDTH_EMU = "9525"
SERIES_LINE_SCHEME_COLORS = ("accent1", "accent2", "accent3", "accent4", "accent5", "accent6")
CHART_COLUMN_STEP = 8
CHART_ROW_STEP = 16
MATRIX_AREA_HEIGHT = 36
DATA_TABLE_GAP_ROWS = 2
DATA_BLOCK_GAP_ROWS = 3
DATA_COMPARE_GAP_COLUMNS = 3

TITLE_FILL = PatternFill("solid", fgColor="1F4E79")
SUBTITLE_FILL = PatternFill("solid", fgColor="D9EAF7")
COMPARE_FILL = PatternFill("solid", fgColor="E2F0D9")
HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


@dataclass(slots=True)
class RadarReportResult:
    output_path: Path
    matrix_count: int
    single_chart_count: int
    compare_chart_count: int
    delta_chart_count: int
    log_lines: list[str] = field(default_factory=list)


@dataclass(frozen=True, slots=True)
class DataTableRef:
    start_row: int
    start_col: int
    header_row: int
    first_data_row: int
    row_count: int
    col_count: int

    @property
    def end_row(self) -> int:
        return self.first_data_row + self.row_count - 1

    @property
    def end_col(self) -> int:
        return self.start_col + self.col_count


def create_radar_report(
    excel_path: str | Path,
    output_dir: str | Path,
    selected_sheets: Iterable[str] | None = None,
) -> Path:
    return generate_radar_report(excel_path, output_dir, selected_sheets).output_path


def generate_radar_report(
    excel_path: str | Path,
    output_dir: str | Path,
    selected_sheets: Iterable[str] | None = None,
    include_delta: bool = False,
) -> RadarReportResult:
    excel_path = Path(excel_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{excel_path.stem}_Radar_Report.xlsx"

    matrices = parse_matrices_from_workbook(excel_path, selected_sheets)
    workbook = Workbook()
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcOnSave = True
    report_ws = workbook.active
    report_ws.title = "Radar_Report"
    data_ws = workbook.create_sheet("Normalized_Data")
    log_ws = workbook.create_sheet("Process_Log")

    chart_counts = {"single": 0, "compare": 0, "delta": 0}
    include_delta = include_delta and _has_multiple_sheets(matrices)
    data_cursor = _prepare_normalized_data(
        data_ws,
        matrices,
        include_compare=_has_multiple_sheets(matrices),
        include_delta=include_delta,
    )
    current_row = 1

    if _has_multiple_sheets(matrices):
        current_row = add_compare_charts(report_ws, data_ws, matrices, current_row, 1, data_cursor, chart_counts)
        if include_delta:
            current_row = add_delta_charts(report_ws, data_ws, matrices, current_row, 1, data_cursor, chart_counts)
    else:
        for matrix in matrices:
            current_row = add_matrix_area(report_ws, data_ws, matrix, current_row, 1, data_cursor, chart_counts)

    log_lines = [
        f"解析矩阵数量: {len(matrices)}",
        f"单图数量: {chart_counts['single']}",
        f"对比图数量: {chart_counts['compare']}",
        f"差值图数量: {chart_counts['delta']}",
        f"输出文件: {output_path}",
        "原始 Excel 只读解析，未修改原始文件。",
        "保存后已原地替换 xl/charts/chart*.xml 为 Excel 原生风格雷达图 XML。",
    ]
    _write_process_log(log_ws, log_lines)

    workbook.save(output_path)
    _replace_saved_radar_charts_with_excel_native_xml(output_path)
    return RadarReportResult(
        output_path=output_path,
        matrix_count=len(matrices),
        single_chart_count=chart_counts["single"],
        compare_chart_count=chart_counts["compare"],
        delta_chart_count=chart_counts["delta"],
        log_lines=log_lines,
    )


def add_matrix_area(
    report_ws: Worksheet,
    data_ws: Worksheet,
    matrix: PatternMatrix,
    start_row: int,
    start_col: int,
    data_cursor: dict[str, object] | None = None,
    chart_counts: dict[str, int] | None = None,
) -> int:
    if data_cursor is None:
        data_cursor = {"row": 1}
    if chart_counts is None:
        chart_counts = {"single": 0, "compare": 0}

    report_ws.cell(row=start_row, column=start_col, value=f"Sheet: {matrix.sheet_name} / Block: {matrix.block_name}")
    add_row_charts(report_ws, data_ws, matrix, start_row + 1, start_col, data_cursor, chart_counts)
    add_col_charts(report_ws, data_ws, matrix, start_row + CHART_ROW_STEP + 1, start_col, data_cursor, chart_counts)
    return start_row + MATRIX_AREA_HEIGHT


def add_row_charts(
    report_ws: Worksheet,
    data_ws: Worksheet,
    matrix: PatternMatrix,
    start_row: int,
    start_col: int,
    data_cursor: dict[str, object],
    chart_counts: dict[str, int],
) -> int:
    table = _get_or_write_matrix_table(data_ws, data_cursor, matrix)
    for index, row_angle in enumerate(matrix.row_angles):
        title = f"{matrix.sheet_name}_{matrix.block_name}_Row_{_angle_text(row_angle)}"
        chart = _create_row_chart_from_matrix_table(data_ws, title, matrix.sheet_name, table, index)
        _place_chart(report_ws, chart, start_row, start_col + index * CHART_COLUMN_STEP)
        chart_counts["single"] += 1
    return start_row + CHART_ROW_STEP


def add_col_charts(
    report_ws: Worksheet,
    data_ws: Worksheet,
    matrix: PatternMatrix,
    start_row: int,
    start_col: int,
    data_cursor: dict[str, object],
    chart_counts: dict[str, int],
) -> int:
    table = _get_or_write_matrix_table(data_ws, data_cursor, matrix)
    for index, col_angle in enumerate(matrix.col_angles):
        title = f"{matrix.sheet_name}_{matrix.block_name}_Col_{_angle_text(col_angle)}"
        chart = _create_col_chart_from_matrix_table(data_ws, title, matrix.sheet_name, table, index)
        _place_chart(report_ws, chart, start_row, start_col + index * CHART_COLUMN_STEP)
        chart_counts["single"] += 1
    return start_row + CHART_ROW_STEP


def add_compare_charts(
    report_ws: Worksheet,
    data_ws: Worksheet,
    matrices: list[PatternMatrix],
    start_row: int,
    start_col: int,
    data_cursor: dict[str, object] | None = None,
    chart_counts: dict[str, int] | None = None,
) -> int:
    if data_cursor is None:
        data_cursor = {"row": 1}
    if chart_counts is None:
        chart_counts = {"single": 0, "compare": 0}

    report_ws.cell(row=start_row, column=start_col, value="Compare Charts")
    _style_report_title(report_ws, start_row, start_col)

    current_row = start_row + 1
    for block_name in _ordered_block_names(matrices):
        block_matrices = [matrix for matrix in matrices if matrix.block_name == block_name]
        report_ws.cell(row=current_row, column=start_col, value=f"Block: {block_name}")
        _style_report_subtitle(report_ws, current_row, start_col)
        _add_compare_row_charts(report_ws, data_ws, block_matrices, current_row + 1, start_col, data_cursor, chart_counts)
        _add_compare_col_charts(
            report_ws,
            data_ws,
            block_matrices,
            current_row + CHART_ROW_STEP + 1,
            start_col,
            data_cursor,
            chart_counts,
        )
        current_row += MATRIX_AREA_HEIGHT
    return current_row


def add_delta_charts(
    report_ws: Worksheet,
    data_ws: Worksheet,
    matrices: list[PatternMatrix],
    start_row: int,
    start_col: int,
    data_cursor: dict[str, object] | None = None,
    chart_counts: dict[str, int] | None = None,
) -> int:
    if data_cursor is None:
        data_cursor = {"row": 1}
    if chart_counts is None:
        chart_counts = {"single": 0, "compare": 0, "delta": 0}
    chart_counts.setdefault("delta", 0)

    report_ws.cell(row=start_row, column=start_col, value="Delta Charts")
    _style_report_title(report_ws, start_row, start_col)

    current_row = start_row + 1
    for block_name in _ordered_block_names(matrices):
        block_matrices = [matrix for matrix in matrices if matrix.block_name == block_name]
        pairs = _delta_matrix_pairs(block_matrices)
        if not pairs:
            continue

        report_ws.cell(row=current_row, column=start_col, value=f"Block: {block_name} / Base: {pairs[0][0].sheet_name}")
        _style_report_subtitle(report_ws, current_row, start_col)
        _add_delta_row_charts(report_ws, data_ws, pairs, current_row + 1, start_col, data_cursor, chart_counts)
        _add_delta_col_charts(
            report_ws,
            data_ws,
            pairs,
            current_row + CHART_ROW_STEP + 1,
            start_col,
            data_cursor,
            chart_counts,
        )
        current_row += MATRIX_AREA_HEIGHT
    return current_row


def _add_compare_row_charts(
    report_ws: Worksheet,
    data_ws: Worksheet,
    matrices: list[PatternMatrix],
    start_row: int,
    start_col: int,
    data_cursor: dict[str, object],
    chart_counts: dict[str, int],
) -> int:
    count = 0
    for block_name in _ordered_block_names(matrices):
        block_matrices = [matrix for matrix in matrices if matrix.block_name == block_name]
        if len({matrix.sheet_name for matrix in block_matrices}) < 2:
            continue
        for row_angle in _common_angles([matrix.row_angles for matrix in block_matrices]):
            title = f"Compare_{block_name}_Row_{_angle_text(row_angle)}"
            table = _get_or_write_compare_table(data_ws, data_cursor, "row", block_name, row_angle, block_matrices)
            chart = _create_chart_from_series_table(data_ws, title, table)
            _place_chart(report_ws, chart, start_row, start_col + count * CHART_COLUMN_STEP)
            chart_counts["compare"] += 1
            count += 1
    return count


def _add_compare_col_charts(
    report_ws: Worksheet,
    data_ws: Worksheet,
    matrices: list[PatternMatrix],
    start_row: int,
    start_col: int,
    data_cursor: dict[str, object],
    chart_counts: dict[str, int],
) -> int:
    count = 0
    for block_name in _ordered_block_names(matrices):
        block_matrices = [matrix for matrix in matrices if matrix.block_name == block_name]
        if len({matrix.sheet_name for matrix in block_matrices}) < 2:
            continue
        for col_angle in _common_angles([matrix.col_angles for matrix in block_matrices]):
            title = f"Compare_{block_name}_Col_{_angle_text(col_angle)}"
            table = _get_or_write_compare_table(data_ws, data_cursor, "col", block_name, col_angle, block_matrices)
            chart = _create_chart_from_series_table(data_ws, title, table)
            _place_chart(report_ws, chart, start_row, start_col + count * CHART_COLUMN_STEP)
            chart_counts["compare"] += 1
            count += 1
    return count


def _add_delta_row_charts(
    report_ws: Worksheet,
    data_ws: Worksheet,
    pairs: list[tuple[PatternMatrix, PatternMatrix]],
    start_row: int,
    start_col: int,
    data_cursor: dict[str, object],
    chart_counts: dict[str, int],
) -> int:
    count = 0
    if not pairs:
        return count
    base_matrix = pairs[0][0]
    for row_angle in _delta_row_angles(pairs):
        title = f"Delta_{base_matrix.block_name}_Row_{_angle_text(row_angle)}"
        table = _get_or_write_delta_table(data_ws, data_cursor, "row", pairs, row_angle)
        chart = _create_chart_from_series_table(data_ws, title, table)
        _place_chart(report_ws, chart, start_row, start_col + count * CHART_COLUMN_STEP)
        chart_counts["delta"] += 1
        count += 1
    return count


def _add_delta_col_charts(
    report_ws: Worksheet,
    data_ws: Worksheet,
    pairs: list[tuple[PatternMatrix, PatternMatrix]],
    start_row: int,
    start_col: int,
    data_cursor: dict[str, object],
    chart_counts: dict[str, int],
) -> int:
    count = 0
    if not pairs:
        return count
    base_matrix = pairs[0][0]
    for col_angle in _delta_col_angles(pairs):
        title = f"Delta_{base_matrix.block_name}_Col_{_angle_text(col_angle)}"
        table = _get_or_write_delta_table(data_ws, data_cursor, "col", pairs, col_angle)
        chart = _create_chart_from_series_table(data_ws, title, table)
        _place_chart(report_ws, chart, start_row, start_col + count * CHART_COLUMN_STEP)
        chart_counts["delta"] += 1
        count += 1
    return count


def _prepare_normalized_data(
    ws: Worksheet,
    matrices: list[PatternMatrix],
    include_compare: bool,
    include_delta: bool,
) -> dict[str, object]:
    layout: dict[str, object] = {
        "row": 1,
        "matrix_tables": {},
        "compare_tables": {},
        "delta_tables": {},
    }
    ws.freeze_panes = "B2"

    current_row = 1
    for block_name in _ordered_block_names(matrices):
        block_matrices = [matrix for matrix in matrices if matrix.block_name == block_name]
        max_source_width = max(
            (max(len(matrix.row_angles), len(matrix.col_angles)) + 1 for matrix in block_matrices),
            default=1,
        )
        source_col = 1
        compare_col = source_col + max_source_width + DATA_COMPARE_GAP_COLUMNS
        delta_col = compare_col + max_source_width + DATA_COMPARE_GAP_COLUMNS

        ws.cell(row=current_row, column=source_col, value=f"Block: {block_name}")
        _style_data_section_title(ws, current_row, source_col, max_source_width)
        source_row = current_row + 1
        for matrix in block_matrices:
            table = _write_matrix_table(ws, matrix, source_row, source_col)
            _matrix_tables(layout)[_matrix_key(matrix)] = table
            source_row = table.end_row + DATA_TABLE_GAP_ROWS

        compare_row = current_row + 1
        if include_compare and len({matrix.sheet_name for matrix in block_matrices}) >= 2:
            ws.cell(row=current_row, column=compare_col, value=f"Compare Data: {block_name}")
            _style_data_section_title(ws, current_row, compare_col, max_source_width, compare=True)
            compare_row = _write_compare_tables_for_block(ws, layout, block_matrices, compare_row, compare_col)

        delta_row = current_row + 1
        if include_delta and _delta_matrix_pairs(block_matrices):
            ws.cell(row=current_row, column=delta_col, value=f"Delta Data: {block_name}")
            _style_data_section_title(ws, current_row, delta_col, max_source_width, compare=True)
            delta_row = _write_delta_tables_for_block(ws, layout, block_matrices, delta_row, delta_col)

        current_row = max(source_row, compare_row, delta_row) + DATA_BLOCK_GAP_ROWS

    layout["row"] = current_row
    _set_normalized_column_widths(ws)
    return layout


def _write_matrix_table(ws: Worksheet, matrix: PatternMatrix, start_row: int, start_col: int) -> DataTableRef:
    title = f"Sheet: {matrix.sheet_name} / Block: {matrix.block_name}"
    header_row = start_row + 1
    first_data_row = start_row + 2

    ws.cell(row=start_row, column=start_col, value=title)
    _style_data_subtitle(ws, start_row, start_col, len(matrix.col_angles) + 1)
    ws.cell(row=header_row, column=start_col, value=f"{matrix.row_label}\\{matrix.col_label}")
    for offset, angle in enumerate(matrix.col_angles, start=1):
        ws.cell(row=header_row, column=start_col + offset, value=_angle_text(angle))

    normalized = matrix.normalized_values()
    for row_index, row_angle in enumerate(matrix.row_angles):
        row = first_data_row + row_index
        ws.cell(row=row, column=start_col, value=_angle_text(row_angle))
        values = normalized[row_index] if row_index < len(normalized) else []
        for col_index in range(len(matrix.col_angles)):
            value = values[col_index] if col_index < len(values) else 0.0
            ws.cell(row=row, column=start_col + col_index + 1, value=value)

    table = DataTableRef(
        start_row=start_row,
        start_col=start_col,
        header_row=header_row,
        first_data_row=first_data_row,
        row_count=len(matrix.row_angles),
        col_count=len(matrix.col_angles),
    )
    _style_data_table(ws, table)
    return table


def _write_series_table(
    ws: Worksheet,
    title: str,
    axes: list[float],
    series: list[tuple[str, list[float | str]]],
    start_row: int,
    start_col: int,
) -> DataTableRef:
    header_row = start_row + 1
    first_data_row = start_row + 2

    ws.cell(row=start_row, column=start_col, value=title)
    _style_data_subtitle(ws, start_row, start_col, len(axes) + 1, compare=True)
    ws.cell(row=header_row, column=start_col, value="Series")
    for offset, angle in enumerate(axes, start=1):
        ws.cell(row=header_row, column=start_col + offset, value=_angle_text(angle))

    for row_index, (series_name, values) in enumerate(series):
        row = first_data_row + row_index
        ws.cell(row=row, column=start_col, value=series_name)
        for col_index in range(len(axes)):
            value = values[col_index] if col_index < len(values) else 0.0
            ws.cell(row=row, column=start_col + col_index + 1, value=value)

    table = DataTableRef(
        start_row=start_row,
        start_col=start_col,
        header_row=header_row,
        first_data_row=first_data_row,
        row_count=len(series),
        col_count=len(axes),
    )
    _style_data_table(ws, table)
    return table


def _write_compare_tables_for_block(
    ws: Worksheet,
    layout: dict[str, object],
    block_matrices: list[PatternMatrix],
    start_row: int,
    start_col: int,
) -> int:
    current_row = start_row
    for row_angle in _common_angles([matrix.row_angles for matrix in block_matrices]):
        title, axes, series = _compare_row_formula_payload(layout, block_matrices, row_angle)
        table = _write_series_table(ws, title, axes, series, current_row, start_col)
        _compare_tables(layout)[_compare_key("row", block_matrices[0].block_name, row_angle)] = table
        current_row = table.end_row + DATA_TABLE_GAP_ROWS

    for col_angle in _common_angles([matrix.col_angles for matrix in block_matrices]):
        title, axes, series = _compare_col_formula_payload(layout, block_matrices, col_angle)
        table = _write_series_table(ws, title, axes, series, current_row, start_col)
        _compare_tables(layout)[_compare_key("col", block_matrices[0].block_name, col_angle)] = table
        current_row = table.end_row + DATA_TABLE_GAP_ROWS

    return current_row


def _write_delta_tables_for_block(
    ws: Worksheet,
    layout: dict[str, object],
    block_matrices: list[PatternMatrix],
    start_row: int,
    start_col: int,
) -> int:
    current_row = start_row
    pairs = _delta_matrix_pairs(block_matrices)
    if not pairs:
        return current_row
    base_matrix = pairs[0][0]

    for row_angle in _delta_row_angles(pairs):
        title, axes, series = _delta_row_formula_payload(layout, pairs, row_angle)
        table = _write_series_table(ws, title, axes, series, current_row, start_col)
        _delta_tables(layout)[_delta_key("row", base_matrix, row_angle)] = table
        current_row = table.end_row + DATA_TABLE_GAP_ROWS

    for col_angle in _delta_col_angles(pairs):
        title, axes, series = _delta_col_formula_payload(layout, pairs, col_angle)
        table = _write_series_table(ws, title, axes, series, current_row, start_col)
        _delta_tables(layout)[_delta_key("col", base_matrix, col_angle)] = table
        current_row = table.end_row + DATA_TABLE_GAP_ROWS

    return current_row


def _get_or_write_matrix_table(
    data_ws: Worksheet,
    data_cursor: dict[str, object],
    matrix: PatternMatrix,
) -> DataTableRef:
    tables = _matrix_tables(data_cursor)
    key = _matrix_key(matrix)
    if key not in tables:
        start_row = int(data_cursor.get("row", 1))
        table = _write_matrix_table(data_ws, matrix, start_row, 1)
        tables[key] = table
        data_cursor["row"] = table.end_row + DATA_TABLE_GAP_ROWS
    return tables[key]


def _get_or_write_compare_table(
    data_ws: Worksheet,
    data_cursor: dict[str, object],
    kind: str,
    block_name: str,
    angle: float,
    block_matrices: list[PatternMatrix],
) -> DataTableRef:
    tables = _compare_tables(data_cursor)
    key = _compare_key(kind, block_name, angle)
    if key not in tables:
        if _has_source_matrix_tables(data_cursor, block_matrices):
            title, axes, series = (
                _compare_row_formula_payload(data_cursor, block_matrices, angle)
                if kind == "row"
                else _compare_col_formula_payload(data_cursor, block_matrices, angle)
            )
        else:
            title, axes, series = (
                _compare_row_payload(block_matrices, angle)
                if kind == "row"
                else _compare_col_payload(block_matrices, angle)
            )
        start_row = int(data_cursor.get("row", 1))
        table = _write_series_table(data_ws, title, axes, series, start_row, 1)
        tables[key] = table
        data_cursor["row"] = table.end_row + DATA_TABLE_GAP_ROWS
    return tables[key]


def _get_or_write_delta_table(
    data_ws: Worksheet,
    data_cursor: dict[str, object],
    kind: str,
    pairs: list[tuple[PatternMatrix, PatternMatrix]],
    angle: float,
) -> DataTableRef:
    base_matrix = pairs[0][0]
    tables = _delta_tables(data_cursor)
    key = _delta_key(kind, base_matrix, angle)
    if key not in tables:
        title, axes, series = (
            _delta_row_formula_payload(data_cursor, pairs, angle)
            if kind == "row"
            else _delta_col_formula_payload(data_cursor, pairs, angle)
        )
        start_row = int(data_cursor.get("row", 1))
        table = _write_series_table(data_ws, title, axes, series, start_row, 1)
        tables[key] = table
        data_cursor["row"] = table.end_row + DATA_TABLE_GAP_ROWS
    return tables[key]


def _create_row_chart_from_matrix_table(
    data_ws: Worksheet,
    title: str,
    series_name: str,
    table: DataTableRef,
    row_index: int,
) -> RadarChart:
    data_row = table.first_data_row + row_index
    chart = _new_radar_chart(title)
    chart.add_data(
        Reference(
            data_ws,
            min_col=table.start_col,
            min_row=data_row,
            max_col=table.end_col,
            max_row=data_row,
        ),
        from_rows=True,
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(data_ws, min_col=table.start_col + 1, min_row=table.header_row, max_col=table.end_col)
    )
    _set_series_titles(chart, [series_name])
    return chart


def _create_col_chart_from_matrix_table(
    data_ws: Worksheet,
    title: str,
    series_name: str,
    table: DataTableRef,
    col_index: int,
) -> RadarChart:
    value_col = table.start_col + col_index + 1
    chart = _new_radar_chart(title)
    chart.add_data(
        Reference(
            data_ws,
            min_col=value_col,
            min_row=table.header_row,
            max_col=value_col,
            max_row=table.end_row,
        ),
        titles_from_data=True,
    )
    chart.set_categories(Reference(data_ws, min_col=table.start_col, min_row=table.first_data_row, max_row=table.end_row))
    _set_series_titles(chart, [series_name])
    return chart


def _create_chart_from_series_table(data_ws: Worksheet, title: str, table: DataTableRef) -> RadarChart:
    chart = _new_radar_chart(title)
    chart.add_data(
        Reference(
            data_ws,
            min_col=table.start_col,
            min_row=table.first_data_row,
            max_col=table.end_col,
            max_row=table.end_row,
        ),
        from_rows=True,
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(data_ws, min_col=table.start_col + 1, min_row=table.header_row, max_col=table.end_col)
    )
    return chart


def _new_radar_chart(title: str) -> RadarChart:
    chart = RadarChart()
    chart.type = "marker"
    chart.style = EXCEL_NATIVE_RADAR_STYLE
    chart.title = title
    chart.width = CHART_WIDTH
    chart.height = CHART_HEIGHT
    return chart


def _set_series_titles(chart: RadarChart, names: list[str]) -> None:
    for series, name in zip(chart.series, names):
        series.tx = SeriesLabel(v=name)


def _place_chart(ws: Worksheet, chart: RadarChart, row: int, column: int) -> None:
    ws.add_chart(chart, f"{get_column_letter(column)}{row}")


def _row_series_for_axes(matrix: PatternMatrix, row_angle: float, axes: list[float]) -> list[float]:
    if row_angle not in matrix.row_angles:
        return [0.0 for _ in axes]
    row_index = matrix.row_angles.index(row_angle)
    normalized = matrix.normalized_values()
    row_values = normalized[row_index] if row_index < len(normalized) else []
    row_by_angle = {
        angle: row_values[index]
        for index, angle in enumerate(matrix.col_angles)
        if index < len(row_values)
    }
    return [row_by_angle.get(angle, 0.0) for angle in axes]


def _col_series_for_axes(matrix: PatternMatrix, col_angle: float, axes: list[float]) -> list[float]:
    if col_angle not in matrix.col_angles:
        return [0.0 for _ in axes]
    col_index = matrix.col_angles.index(col_angle)
    col_values = matrix.col_values(col_index)
    col_by_angle = {
        angle: col_values[index]
        for index, angle in enumerate(matrix.row_angles)
        if index < len(col_values)
    }
    return [col_by_angle.get(angle, 0.0) for angle in axes]


def _compare_row_payload(
    block_matrices: list[PatternMatrix],
    row_angle: float,
) -> tuple[str, list[float], list[tuple[str, list[float]]]]:
    block_name = block_matrices[0].block_name
    axes = _union_angles([matrix.col_angles for matrix in block_matrices])
    series = [(matrix.sheet_name, _row_series_for_axes(matrix, row_angle, axes)) for matrix in block_matrices]
    return f"Compare_{block_name}_Row_{_angle_text(row_angle)}", axes, series


def _compare_col_payload(
    block_matrices: list[PatternMatrix],
    col_angle: float,
) -> tuple[str, list[float], list[tuple[str, list[float]]]]:
    block_name = block_matrices[0].block_name
    axes = _union_angles([matrix.row_angles for matrix in block_matrices])
    series = [(matrix.sheet_name, _col_series_for_axes(matrix, col_angle, axes)) for matrix in block_matrices]
    return f"Compare_{block_name}_Col_{_angle_text(col_angle)}", axes, series


def _compare_row_formula_payload(
    layout: dict[str, object],
    block_matrices: list[PatternMatrix],
    row_angle: float,
) -> tuple[str, list[float], list[tuple[str, list[float | str]]]]:
    block_name = block_matrices[0].block_name
    axes = _union_angles([matrix.col_angles for matrix in block_matrices])
    series = [
        (matrix.sheet_name, [_row_formula_or_zero(layout, matrix, row_angle, col_angle) for col_angle in axes])
        for matrix in block_matrices
    ]
    return f"Compare_{block_name}_Row_{_angle_text(row_angle)}", axes, series


def _compare_col_formula_payload(
    layout: dict[str, object],
    block_matrices: list[PatternMatrix],
    col_angle: float,
) -> tuple[str, list[float], list[tuple[str, list[float | str]]]]:
    block_name = block_matrices[0].block_name
    axes = _union_angles([matrix.row_angles for matrix in block_matrices])
    series = [
        (matrix.sheet_name, [_col_formula_or_zero(layout, matrix, col_angle, row_angle) for row_angle in axes])
        for matrix in block_matrices
    ]
    return f"Compare_{block_name}_Col_{_angle_text(col_angle)}", axes, series


def _delta_row_formula_payload(
    layout: dict[str, object],
    pairs: list[tuple[PatternMatrix, PatternMatrix]],
    row_angle: float,
) -> tuple[str, list[float], list[tuple[str, list[float | str]]]]:
    base_matrix = pairs[0][0]
    active_pairs = [
        (base_matrix, target_matrix)
        for base_matrix, target_matrix in pairs
        if row_angle in base_matrix.row_angles and row_angle in target_matrix.row_angles
    ]
    axes = _union_common_target_angles(active_pairs, axis="col")
    series = [
        (
            f"{target_matrix.sheet_name} - {base_matrix.sheet_name}",
            [_delta_formula_or_zero(layout, base_matrix, target_matrix, row_angle, col_angle) for col_angle in axes],
        )
        for base_matrix, target_matrix in active_pairs
    ]
    return (
        f"Delta_{base_matrix.block_name}_Row_{_angle_text(row_angle)}",
        axes,
        series,
    )


def _delta_col_formula_payload(
    layout: dict[str, object],
    pairs: list[tuple[PatternMatrix, PatternMatrix]],
    col_angle: float,
) -> tuple[str, list[float], list[tuple[str, list[float | str]]]]:
    base_matrix = pairs[0][0]
    active_pairs = [
        (base_matrix, target_matrix)
        for base_matrix, target_matrix in pairs
        if col_angle in base_matrix.col_angles and col_angle in target_matrix.col_angles
    ]
    axes = _union_common_target_angles(active_pairs, axis="row")
    series = [
        (
            f"{target_matrix.sheet_name} - {base_matrix.sheet_name}",
            [_delta_formula_or_zero(layout, base_matrix, target_matrix, row_angle, col_angle) for row_angle in axes],
        )
        for base_matrix, target_matrix in active_pairs
    ]
    return (
        f"Delta_{base_matrix.block_name}_Col_{_angle_text(col_angle)}",
        axes,
        series,
    )


def _row_formula_or_zero(
    layout: dict[str, object],
    matrix: PatternMatrix,
    row_angle: float,
    col_angle: float,
) -> float | str:
    if row_angle not in matrix.row_angles or col_angle not in matrix.col_angles:
        return 0.0
    table = _matrix_tables(layout).get(_matrix_key(matrix))
    if table is None:
        return _row_series_for_axes(matrix, row_angle, [col_angle])[0]
    row = table.first_data_row + matrix.row_angles.index(row_angle)
    col = table.start_col + matrix.col_angles.index(col_angle) + 1
    return _cell_formula(row, col)


def _delta_formula_or_zero(
    layout: dict[str, object],
    base_matrix: PatternMatrix,
    target_matrix: PatternMatrix,
    row_angle: float,
    col_angle: float,
) -> float | str:
    if (
        row_angle not in base_matrix.row_angles
        or row_angle not in target_matrix.row_angles
        or col_angle not in base_matrix.col_angles
        or col_angle not in target_matrix.col_angles
    ):
        return 0.0

    base_ref = _matrix_cell_reference(layout, base_matrix, row_angle, col_angle)
    target_ref = _matrix_cell_reference(layout, target_matrix, row_angle, col_angle)
    if base_ref is not None and target_ref is not None:
        return f"={target_ref}-{base_ref}"

    target_value = _row_series_for_axes(target_matrix, row_angle, [col_angle])[0]
    base_value = _row_series_for_axes(base_matrix, row_angle, [col_angle])[0]
    return target_value - base_value


def _col_formula_or_zero(
    layout: dict[str, object],
    matrix: PatternMatrix,
    col_angle: float,
    row_angle: float,
) -> float | str:
    if col_angle not in matrix.col_angles or row_angle not in matrix.row_angles:
        return 0.0
    table = _matrix_tables(layout).get(_matrix_key(matrix))
    if table is None:
        return _col_series_for_axes(matrix, col_angle, [row_angle])[0]
    row = table.first_data_row + matrix.row_angles.index(row_angle)
    col = table.start_col + matrix.col_angles.index(col_angle) + 1
    return _cell_formula(row, col)


def _matrix_cell_reference(
    layout: dict[str, object],
    matrix: PatternMatrix,
    row_angle: float,
    col_angle: float,
) -> str | None:
    table = _matrix_tables(layout).get(_matrix_key(matrix))
    if table is None or row_angle not in matrix.row_angles or col_angle not in matrix.col_angles:
        return None
    row = table.first_data_row + matrix.row_angles.index(row_angle)
    col = table.start_col + matrix.col_angles.index(col_angle) + 1
    return _cell_reference(row, col)


def _ordered_block_names(matrices: list[PatternMatrix]) -> list[str]:
    names: list[str] = []
    for matrix in matrices:
        if matrix.block_name not in names:
            names.append(matrix.block_name)
    return names


def _common_angles(angle_lists: list[list[float]]) -> list[float]:
    counts: dict[float, int] = {}
    for angles in angle_lists:
        for angle in set(angles):
            counts[angle] = counts.get(angle, 0) + 1
    return sorted(angle for angle, count in counts.items() if count >= 2)


def _union_angles(angle_lists: list[list[float]]) -> list[float]:
    return sorted({angle for angles in angle_lists for angle in angles})


def _has_multiple_sheets(matrices: list[PatternMatrix]) -> bool:
    return len({matrix.sheet_name for matrix in matrices}) > 1


def _delta_matrix_pairs(block_matrices: list[PatternMatrix]) -> list[tuple[PatternMatrix, PatternMatrix]]:
    if len({matrix.sheet_name for matrix in block_matrices}) < 2:
        return []
    base_matrix = block_matrices[0]
    return [(base_matrix, matrix) for matrix in block_matrices[1:] if matrix.sheet_name != base_matrix.sheet_name]


def _delta_row_angles(pairs: list[tuple[PatternMatrix, PatternMatrix]]) -> list[float]:
    return sorted(
        {
            row_angle
            for base_matrix, target_matrix in pairs
            for row_angle in _common_angles([base_matrix.row_angles, target_matrix.row_angles])
            if _common_angles([base_matrix.col_angles, target_matrix.col_angles])
        }
    )


def _delta_col_angles(pairs: list[tuple[PatternMatrix, PatternMatrix]]) -> list[float]:
    return sorted(
        {
            col_angle
            for base_matrix, target_matrix in pairs
            for col_angle in _common_angles([base_matrix.col_angles, target_matrix.col_angles])
            if _common_angles([base_matrix.row_angles, target_matrix.row_angles])
        }
    )


def _union_common_target_angles(pairs: list[tuple[PatternMatrix, PatternMatrix]], axis: str) -> list[float]:
    angle_sets = [
        _common_angles(
            [
                base_matrix.row_angles if axis == "row" else base_matrix.col_angles,
                target_matrix.row_angles if axis == "row" else target_matrix.col_angles,
            ]
        )
        for base_matrix, target_matrix in pairs
    ]
    return _union_angles(angle_sets)


def _write_process_log(ws: Worksheet, lines: list[str]) -> None:
    ws.cell(row=1, column=1, value="Process Log")
    for index, line in enumerate(lines, start=2):
        ws.cell(row=index, column=1, value=line)


def _angle_text(angle: float) -> str:
    if float(angle).is_integer():
        return str(int(angle))
    return f"{angle:.10f}".rstrip("0").rstrip(".")


def _matrix_key(matrix: PatternMatrix) -> tuple[int, str, str]:
    return (id(matrix), matrix.sheet_name, matrix.block_name)


def _compare_key(kind: str, block_name: str, angle: float) -> tuple[str, str, float]:
    return kind, block_name, angle


def _delta_key(
    kind: str,
    base_matrix: PatternMatrix,
    angle: float,
) -> tuple[str, str, str, float]:
    return kind, base_matrix.sheet_name, base_matrix.block_name, angle


def _matrix_tables(layout: dict[str, object]) -> dict[tuple[int, str, str], DataTableRef]:
    return layout.setdefault("matrix_tables", {})  # type: ignore[return-value]


def _compare_tables(layout: dict[str, object]) -> dict[tuple[str, str, float], DataTableRef]:
    return layout.setdefault("compare_tables", {})  # type: ignore[return-value]


def _delta_tables(layout: dict[str, object]) -> dict[tuple[str, str, str, float], DataTableRef]:
    return layout.setdefault("delta_tables", {})  # type: ignore[return-value]


def _has_source_matrix_tables(layout: dict[str, object], matrices: list[PatternMatrix]) -> bool:
    tables = _matrix_tables(layout)
    return all(_matrix_key(matrix) in tables for matrix in matrices)


def _cell_formula(row: int, col: int) -> str:
    return f"={_cell_reference(row, col)}"


def _cell_reference(row: int, col: int) -> str:
    return f"${get_column_letter(col)}${row}"


def _style_report_title(ws: Worksheet, row: int, col: int) -> None:
    cell = ws.cell(row=row, column=col)
    cell.font = Font(bold=True, size=14)


def _style_report_subtitle(ws: Worksheet, row: int, col: int) -> None:
    cell = ws.cell(row=row, column=col)
    cell.font = Font(bold=True, size=12)


def _style_data_section_title(
    ws: Worksheet,
    row: int,
    col: int,
    width: int,
    compare: bool = False,
) -> None:
    for column in range(col, col + max(width, 1)):
        cell = ws.cell(row=row, column=column)
        cell.fill = COMPARE_FILL if compare else TITLE_FILL
        cell.font = Font(bold=True, color="000000" if compare else "FFFFFF")
        cell.alignment = Alignment(horizontal="center")


def _style_data_subtitle(
    ws: Worksheet,
    row: int,
    col: int,
    width: int,
    compare: bool = False,
) -> None:
    for column in range(col, col + max(width, 1)):
        cell = ws.cell(row=row, column=column)
        cell.fill = COMPARE_FILL if compare else SUBTITLE_FILL
        cell.font = Font(bold=True)


def _style_data_table(ws: Worksheet, table: DataTableRef) -> None:
    for row in range(table.header_row, table.end_row + 1):
        for col in range(table.start_col, table.end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if row == table.header_row:
                cell.fill = HEADER_FILL
                cell.font = Font(bold=True)


def _set_normalized_column_widths(ws: Worksheet) -> None:
    for column in range(1, ws.max_column + 1):
        letter = get_column_letter(column)
        ws.column_dimensions[letter].width = 12
    ws.column_dimensions["A"].width = 22


_CHART_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"
_DRAWING_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
_CHART_PART_PREFIX = "xl/charts/chart"
_CHART_PART_SUFFIX = ".xml"

ET.register_namespace("c", _CHART_NS)
ET.register_namespace("a", _DRAWING_NS)


def _replace_saved_radar_charts_with_excel_native_xml(xlsx_path: Path) -> None:
    """Rewrite saved radar chart parts with an Excel-like XML shape.

    openpyxl remains responsible for workbook structure, sheets, drawings,
    relationships, and chart data references.  After saving, this function
    performs an in-place ZIP package rewrite of every generated radar chart
    part so the chart XML contains the default elements that Excel writes for
    native radar charts but openpyxl omits.
    """
    tmp_path: Path | None = None
    payloads: list[tuple[zipfile.ZipInfo, bytes]] = []
    try:
        with zipfile.ZipFile(xlsx_path, "r") as source:
            parts = {info.filename: info for info in source.infolist()}
            chart_names = [name for name in parts if _is_chart_part(name)]
            if not chart_names:
                return

            replacements: dict[str, bytes] = {}
            for chart_name in chart_names:
                chart_xml = source.read(chart_name)
                replacement = _excel_native_radar_chart_xml(chart_xml)
                if replacement is not None:
                    replacements[chart_name] = replacement

            if not replacements:
                return

            for info in source.infolist():
                payloads.append((info, replacements.get(info.filename, source.read(info.filename))))

        with NamedTemporaryFile(delete=False, dir=xlsx_path.parent, suffix=".xlsx") as tmp_file:
            tmp_path = Path(tmp_file.name)

        with zipfile.ZipFile(tmp_path, "w") as target:
            for info, data in payloads:
                target.writestr(info, data)

        tmp_path.replace(xlsx_path)
    except Exception:
        if tmp_path is not None:
            tmp_path.unlink(missing_ok=True)
        raise


def _is_chart_part(name: str) -> bool:
    chart_number = name.removeprefix(_CHART_PART_PREFIX).removesuffix(_CHART_PART_SUFFIX)
    return (
        name.startswith(_CHART_PART_PREFIX)
        and name.endswith(_CHART_PART_SUFFIX)
        and chart_number.isdigit()
    )


def _excel_native_radar_chart_xml(chart_xml: bytes) -> bytes | None:
    root = ET.fromstring(chart_xml)
    chart = root.find(_c("chart"))
    plot_area = chart.find(_c("plotArea")) if chart is not None else None
    radar_chart = plot_area.find(_c("radarChart")) if plot_area is not None else None
    if chart is None or plot_area is None or radar_chart is None:
        return None

    _ensure_child(root, "date1904", before={"style", "chart"}, val="0")
    _ensure_child(root, "lang", before={"style", "chart"}, val="zh-CN")
    _ensure_child(root, "roundedCorners", before={"style", "chart"}, val="0")
    _ensure_child(root, "style", before={"chart"}, val=str(EXCEL_NATIVE_RADAR_STYLE))

    title = chart.find(_c("title"))
    if title is not None:
        _ensure_child(title, "overlay", before={"spPr", "txPr"}, val="0")
        _set_no_fill_shape(_ensure_child(title, "spPr", before={"txPr"}))
        _normalise_text_properties(title, font_size="1400", rotate=False)
    _ensure_child(chart, "autoTitleDeleted", before={"plotArea"}, val="0")

    _ensure_child(plot_area, "layout", before={"radarChart", "catAx", "valAx"})
    _ensure_child(radar_chart, "radarStyle", before={"varyColors", "varyingColors", "ser", "dLbls", "axId"}, val="marker")
    _remove_child(radar_chart, "varyingColors")
    _ensure_child(radar_chart, "varyColors", before={"ser", "dLbls", "axId"}, val="0")
    for series in radar_chart.findall(_c("ser")):
        _normalise_radar_series(series)
    _normalise_data_labels(radar_chart)

    for axis_name in ("catAx", "valAx"):
        for axis in plot_area.findall(_c(axis_name)):
            _normalise_axis(axis, value_axis=(axis_name == "valAx"))

    legend = chart.find(_c("legend"))
    if len(radar_chart.findall(_c("ser"))) <= 1:
        if legend is not None:
            chart.remove(legend)
    elif legend is not None:
        _ensure_child(legend, "legendPos", before={"layout", "overlay", "spPr", "txPr"}, val="r")
        _ensure_child(legend, "overlay", before={"spPr", "txPr"}, val="0")
        _set_no_fill_shape(_ensure_child(legend, "spPr", before={"txPr"}))
        _normalise_text_properties(legend, font_size="900", rotate=False)

    _ensure_child(chart, "plotVisOnly", before={"dispBlanksAs", "showDLblsOverMax"}, val="1")
    _ensure_child(chart, "dispBlanksAs", before={"showDLblsOverMax"}, val="gap")
    _ensure_child(chart, "showDLblsOverMax", val="0")
    _set_no_fill_shape(_ensure_child(plot_area, "spPr"))
    _normalise_chart_space(root)
    _ensure_print_settings(root)
    return ET.tostring(root, encoding="utf-8", xml_declaration=False)


def _normalise_radar_series(series: ET.Element) -> None:
    sp_pr = _ensure_child(series, "spPr", before={"marker", "dPt", "dLbls", "cat", "val"})
    _normalise_series_line(sp_pr, _series_scheme_color(series))
    marker = _ensure_child(series, "marker", before={"dPt", "dLbls", "cat", "val"})
    _ensure_child(marker, "symbol", before={"size", "spPr"}, val="none")
    _remove_child(marker, "size")
    _remove_child(marker, "spPr")


def _normalise_series_line(sp_pr: ET.Element, scheme_color: str) -> None:
    _clear_element(sp_pr)
    line = sp_pr.find(_a("ln"))
    if line is None:
        line = ET.SubElement(sp_pr, _a("ln"))
    line.set("w", SERIES_LINE_WIDTH_EMU)
    line.set("cap", "rnd")
    solid_fill = ET.SubElement(line, _a("solidFill"))
    ET.SubElement(solid_fill, _a("schemeClr"), {"val": scheme_color})
    ET.SubElement(line, _a("round"))
    ET.SubElement(sp_pr, _a("effectLst"))


def _series_scheme_color(series: ET.Element) -> str:
    idx = series.find(_c("idx"))
    try:
        color_index = int(idx.get("val", "0")) if idx is not None else 0
    except ValueError:
        color_index = 0
    return SERIES_LINE_SCHEME_COLORS[color_index % len(SERIES_LINE_SCHEME_COLORS)]


def _normalise_data_labels(radar_chart: ET.Element) -> None:
    data_labels = _ensure_child(radar_chart, "dLbls", before={"axId"})
    for tag_name in ("showLegendKey", "showVal", "showCatName", "showSerName", "showPercent", "showBubbleSize"):
        _ensure_child(data_labels, tag_name, val="0")


def _normalise_axis(axis: ET.Element, value_axis: bool) -> None:
    _ensure_child(axis, "delete", before={"axPos", "numFmt", "majorGridlines", "majorTickMark"}, val="0")
    _ensure_child(axis, "axPos", before={"numFmt", "majorGridlines", "majorTickMark"}, val="l" if value_axis else "b")
    _ensure_child(
        axis,
        "numFmt",
        before={"majorGridlines", "majorTickMark", "minorTickMark"},
        formatCode="General",
        sourceLinked="1",
    )
    _ensure_child(axis, "majorTickMark", before={"minorTickMark", "tickLblPos", "spPr", "txPr"}, val="none")
    _ensure_child(axis, "minorTickMark", before={"tickLblPos", "spPr", "txPr"}, val="none")
    _ensure_child(axis, "tickLblPos", before={"spPr", "txPr", "crossAx", "crosses"}, val="nextTo")
    if value_axis:
        major_gridlines = _ensure_child(axis, "majorGridlines", before={"numFmt", "majorTickMark"})
        _set_light_line_shape(_ensure_child(major_gridlines, "spPr"))
        _set_no_fill_shape(_ensure_child(axis, "spPr", before={"txPr", "crossAx", "crosses"}))
    else:
        _set_light_line_shape(_ensure_child(axis, "spPr", before={"txPr", "crossAx", "crosses"}), no_fill=True)
    _normalise_text_properties(axis, font_size="900", rotate=True)
    if value_axis:
        _ensure_child(axis, "crosses", before={"crossBetween"}, val="autoZero")
        _ensure_child(axis, "crossBetween", val="between")
    else:
        _ensure_child(axis, "crosses", before={"auto", "lblAlgn", "lblOffset", "noMultiLvlLbl"}, val="autoZero")
        _ensure_child(axis, "auto", before={"lblAlgn", "lblOffset", "noMultiLvlLbl"}, val="1")
        _ensure_child(axis, "lblAlgn", before={"lblOffset", "noMultiLvlLbl"}, val="ctr")
        _ensure_child(axis, "lblOffset", before={"noMultiLvlLbl"}, val="100")
        _ensure_child(axis, "noMultiLvlLbl", val="0")


def _ensure_print_settings(root: ET.Element) -> None:
    print_settings = _ensure_child(root, "printSettings")
    _ensure_child(print_settings, "headerFooter")
    margins = _ensure_child(print_settings, "pageMargins")
    margin_defaults = {
        "l": "0.7",
        "r": "0.7",
        "t": "0.75",
        "b": "0.75",
        "header": "0.3",
        "footer": "0.3",
    }
    for attr, value in margin_defaults.items():
        margins.set(attr, value)
    _ensure_child(print_settings, "pageSetup")


def _ensure_shape_properties(parent: ET.Element) -> ET.Element:
    sp_pr = _ensure_child(parent, "spPr", before={"txPr", "crossAx", "crosses"})
    _ensure_line(sp_pr)
    return sp_pr


def _ensure_line(sp_pr: ET.Element) -> ET.Element:
    line = sp_pr.find(_a("ln"))
    if line is None:
        line = ET.SubElement(sp_pr, _a("ln"))
    if line.find(_a("prstDash")) is None:
        ET.SubElement(line, _a("prstDash"), {"val": "solid"})
    return line


def _ensure_text_properties(parent: ET.Element) -> ET.Element:
    tx_pr = _ensure_child(parent, "txPr", before={"crossAx", "crosses"})
    if tx_pr.find(_a("bodyPr")) is None:
        ET.SubElement(tx_pr, _a("bodyPr"))
    if tx_pr.find(_a("lstStyle")) is None:
        ET.SubElement(tx_pr, _a("lstStyle"))
    if tx_pr.find(_a("p")) is None:
        paragraph = ET.SubElement(tx_pr, _a("p"))
        paragraph_properties = ET.SubElement(paragraph, _a("pPr"))
        ET.SubElement(paragraph_properties, _a("defRPr"))
    return tx_pr


def _normalise_chart_space(root: ET.Element) -> None:
    sp_pr = _ensure_child(root, "spPr", before={"txPr", "printSettings"})
    _clear_element(sp_pr)
    solid_fill = ET.SubElement(sp_pr, _a("solidFill"))
    ET.SubElement(solid_fill, _a("schemeClr"), {"val": "bg1"})
    _append_light_line(sp_pr)
    ET.SubElement(sp_pr, _a("effectLst"))
    _normalise_text_properties(root, font_size=None, rotate=False)


def _set_no_fill_shape(sp_pr: ET.Element) -> None:
    _clear_element(sp_pr)
    ET.SubElement(sp_pr, _a("noFill"))
    line = ET.SubElement(sp_pr, _a("ln"))
    ET.SubElement(line, _a("noFill"))
    ET.SubElement(sp_pr, _a("effectLst"))


def _set_light_line_shape(sp_pr: ET.Element, no_fill: bool = False) -> None:
    _clear_element(sp_pr)
    if no_fill:
        ET.SubElement(sp_pr, _a("noFill"))
    _append_light_line(sp_pr)
    ET.SubElement(sp_pr, _a("effectLst"))


def _append_light_line(parent: ET.Element) -> ET.Element:
    line = ET.SubElement(
        parent,
        _a("ln"),
        {"w": GRID_LINE_WIDTH_EMU, "cap": "flat", "cmpd": "sng", "algn": "ctr"},
    )
    solid_fill = ET.SubElement(line, _a("solidFill"))
    scheme_color = ET.SubElement(solid_fill, _a("schemeClr"), {"val": "tx1"})
    ET.SubElement(scheme_color, _a("lumMod"), {"val": "15000"})
    ET.SubElement(scheme_color, _a("lumOff"), {"val": "85000"})
    ET.SubElement(line, _a("round"))
    return line


def _normalise_text_properties(parent: ET.Element, font_size: str | None, rotate: bool) -> ET.Element:
    tx_pr = _ensure_child(parent, "txPr", before={"crossAx", "crosses", "printSettings"})
    _clear_element(tx_pr)
    body_attrs = {
        "spcFirstLastPara": "1",
        "vertOverflow": "ellipsis",
        "vert": "horz",
        "wrap": "square",
        "anchor": "ctr",
        "anchorCtr": "1",
    }
    if rotate:
        body_attrs["rot"] = "-60000000"
    elif font_size is not None:
        body_attrs["rot"] = "0"
    ET.SubElement(tx_pr, _a("bodyPr"), body_attrs)
    ET.SubElement(tx_pr, _a("lstStyle"))
    paragraph = ET.SubElement(tx_pr, _a("p"))
    paragraph_properties = ET.SubElement(paragraph, _a("pPr"))
    run_attrs = {
        "b": "0",
        "i": "0",
        "u": "none",
        "strike": "noStrike",
        "kern": "1200",
        "baseline": "0",
    }
    if font_size is not None:
        run_attrs["sz"] = font_size
    if font_size == "1400":
        run_attrs["spc"] = "0"
    default_run = ET.SubElement(paragraph_properties, _a("defRPr"), run_attrs)
    if font_size is not None:
        solid_fill = ET.SubElement(default_run, _a("solidFill"))
        scheme_color = ET.SubElement(solid_fill, _a("schemeClr"), {"val": "tx1"})
        ET.SubElement(scheme_color, _a("lumMod"), {"val": "65000"})
        ET.SubElement(scheme_color, _a("lumOff"), {"val": "35000"})
        ET.SubElement(default_run, _a("latin"), {"typeface": "+mn-lt"})
        ET.SubElement(default_run, _a("ea"), {"typeface": "+mn-ea"})
        ET.SubElement(default_run, _a("cs"), {"typeface": "+mn-cs"})
    ET.SubElement(paragraph, _a("endParaRPr"), {"lang": "zh-CN"})
    return tx_pr


def _clear_element(element: ET.Element) -> None:
    element.attrib.clear()
    for child in list(element):
        element.remove(child)


def _ensure_child(
    parent: ET.Element,
    tag_name: str,
    before: set[str] | None = None,
    **attributes: str,
) -> ET.Element:
    tag = _c(tag_name)
    child = parent.find(tag)
    if child is None:
        child = ET.Element(tag)
        insert_at = _insertion_index(parent, before or set())
        parent.insert(insert_at, child)
    for name, value in attributes.items():
        child.set(name, value)
    return child


def _remove_child(parent: ET.Element, tag_name: str) -> None:
    child = parent.find(_c(tag_name))
    if child is not None:
        parent.remove(child)


def _insertion_index(parent: ET.Element, before: set[str]) -> int:
    if not before:
        return len(parent)
    before_tags = {_c(tag) for tag in before}
    for index, child in enumerate(list(parent)):
        if child.tag in before_tags:
            return index
    return len(parent)


def _c(tag_name: str) -> str:
    return f"{{{_CHART_NS}}}{tag_name}"


def _a(tag_name: str) -> str:
    return f"{{{_DRAWING_NS}}}{tag_name}"
