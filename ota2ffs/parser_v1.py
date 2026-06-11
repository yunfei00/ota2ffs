from __future__ import annotations

from typing import TypeAlias

from openpyxl.worksheet.worksheet import Worksheet

from .utils import (
    FarFieldSource,
    complete_angle_range,
    contains_text,
    infer_step,
    is_text,
    normalize_angle,
    sorted_unique,
    to_float,
)


BlockStart: TypeAlias = tuple[int, int]


def is_v1_sheet(ws: Worksheet) -> bool:
    return _find_block(ws, "Theta") is not None and _find_block(ws, "Phi") is not None


def parse_sources(ws: Worksheet) -> list[FarFieldSource]:
    sources = [parse_sheet(ws)]
    total_start = _find_block(ws, "Total")
    if total_start is not None:
        total_headers, total_phi_values, total_raw = _parse_block(ws, total_start)
        total_values = {
            (normalize_angle(raw_phi + 180), normalize_angle(theta)): value
            for (raw_phi, theta), value in total_raw.items()
        }
        sources.append(
            FarFieldSource(
                sheet_name=ws.title,
                suffix="_total",
                theta_angles=_complete_theta_angles(total_headers),
                phi_angles=_complete_phi_angles(total_phi_values),
                e_theta_db=total_values,
                e_phi_db={},
                version="V1",
            )
        )
    return sources


def parse_sheet(ws: Worksheet) -> FarFieldSource:
    theta_start = _find_block(ws, "Theta")
    phi_start = _find_block(ws, "Phi")
    if theta_start is None or phi_start is None:
        raise ValueError("未找到 V1 所需的 Theta/Phi 数据块")

    theta_headers, theta_phi_values, e_theta_raw = _parse_block(ws, theta_start)
    phi_headers, phi_phi_values, e_phi_raw = _parse_block(ws, phi_start)

    theta_angles = _complete_theta_angles([*theta_headers, *phi_headers])
    phi_angles = _complete_phi_angles([*theta_phi_values, *phi_phi_values])

    e_theta = {
        (normalize_angle(raw_phi + 180), normalize_angle(theta)): value
        for (raw_phi, theta), value in e_theta_raw.items()
    }
    e_phi = {
        (normalize_angle(raw_phi + 180), normalize_angle(theta)): value
        for (raw_phi, theta), value in e_phi_raw.items()
    }

    return FarFieldSource(
        sheet_name=ws.title,
        theta_angles=theta_angles,
        phi_angles=phi_angles,
        e_theta_db=e_theta,
        e_phi_db=e_phi,
        version="V1",
    )


def _find_block(ws: Worksheet, title: str) -> BlockStart | None:
    for row in range(1, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
            if (
                is_text(ws.cell(row=row, column=column).value, title)
                and contains_text(ws.cell(row=row, column=column + 1).value, "Phi Angle")
                and contains_text(ws.cell(row=row + 1, column=column + 1).value, "Theta Angle")
            ):
                return row, column
    return None


def _parse_block(ws: Worksheet, start: BlockStart) -> tuple[list[float], list[float], dict[tuple[float, float], float]]:
    start_row, start_column = start
    angle_column = start_column + 1
    theta_columns = _read_theta_columns(ws, start_row, start_column + 2)
    if not theta_columns:
        raise ValueError(f"第 {start_row} 行未找到 Theta 角度")

    theta_angles = [theta for _, theta in theta_columns]
    phi_angles: list[float] = []
    values: dict[tuple[float, float], float] = {}
    saw_data = False

    for row in range(start_row + 2, ws.max_row + 1):
        raw_phi = to_float(ws.cell(row=row, column=angle_column).value)
        if raw_phi is None:
            if saw_data:
                break
            continue

        raw_phi = normalize_angle(raw_phi)
        phi_angles.append(raw_phi)
        saw_data = True

        for column, theta in theta_columns:
            db_value = to_float(ws.cell(row=row, column=column).value)
            if db_value is not None:
                values[(raw_phi, theta)] = db_value

    if not phi_angles:
        raise ValueError(f"第 {start_row} 行数据块未找到 Phi 角度")

    return theta_angles, phi_angles, values


def _read_theta_columns(ws: Worksheet, start_row: int, first_column: int) -> list[tuple[int, float]]:
    theta_columns: list[tuple[int, float]] = []
    for column in range(first_column, ws.max_column + 1):
        theta = to_float(ws.cell(row=start_row, column=column).value)
        if theta is None:
            if theta_columns:
                break
            continue
        theta_columns.append((column, normalize_angle(theta)))
    return theta_columns


def _complete_theta_angles(actual_angles: list[float]) -> list[float]:
    step = infer_step(actual_angles)
    if step is None:
        return sorted_unique(actual_angles)
    return complete_angle_range(0, 180, step)


def _complete_phi_angles(raw_phi_angles: list[float]) -> list[float]:
    output_angles = [normalize_angle(phi + 180) for phi in raw_phi_angles]
    step = infer_step(output_angles)
    if step is None:
        return sorted_unique(output_angles)
    return complete_angle_range(0, 360, step)
