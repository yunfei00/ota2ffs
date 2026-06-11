from __future__ import annotations

import traceback
from collections.abc import Callable
from pathlib import Path

from openpyxl import load_workbook
from PySide6.QtCore import QObject, QSettings, QThread, Qt, QUrl, Signal, Slot
from PySide6.QtGui import QDesktopServices
from PySide6.QtWidgets import (
    QApplication,
    QCheckBox,
    QComboBox,
    QFileDialog,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from .converter import convert_excel
from .radar_report import generate_radar_report
from .utils import FREQUENCY_UNITS, get_default_output_dir, sanitize_filename


SETTINGS_ORG = "ota2ffs"
SETTINGS_APP = "OTA2FFS Converter"
LAST_EXCEL_KEY = "last_excel_path"
SHEET_CHECKED_PREFIX = "✓  "
SHEET_UNCHECKED_PREFIX = "    "
MIN_VISIBLE_SHEET_ROWS = 5
MAX_VISIBLE_SHEET_ROWS = 14


class _TaskWorker(QObject):
    succeeded = Signal(object)
    failed = Signal(str)
    done = Signal()

    def __init__(self, task: Callable[[], object]) -> None:
        super().__init__()
        self._task = task

    @Slot()
    def run(self) -> None:
        try:
            self.succeeded.emit(self._task())
        except Exception:
            self.failed.emit(traceback.format_exc())
        finally:
            self.done.emit()


class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("OTA2FFS Converter")
        self.resize(920, 720)
        self.settings = QSettings(SETTINGS_ORG, SETTINGS_APP)
        self.excel_path: Path | None = None
        self.output_dir: Path | None = get_default_output_dir()
        self.last_output_dir: Path | None = self.output_dir
        self._updating_sheet_item = False
        self._sheet_selection_order: list[str] = []
        self._task_thread: QThread | None = None
        self._task_worker: _TaskWorker | None = None
        self._task_success_handler: Callable[[object], None] | None = None
        self._task_error_title = ""
        self._build_ui()
        self._apply_styles()
        self._load_last_excel()

    def _build_ui(self) -> None:
        central = QWidget()
        central.setObjectName("appRoot")
        root = QVBoxLayout(central)
        root.setContentsMargins(18, 18, 18, 18)
        root.setSpacing(14)

        path_group = QGroupBox("输入与输出")
        path_layout = QGridLayout(path_group)
        path_layout.setContentsMargins(16, 20, 16, 16)
        path_layout.setHorizontalSpacing(10)
        path_layout.setVerticalSpacing(12)

        self.excel_edit = QLineEdit()
        self.excel_edit.setReadOnly(True)
        self.choose_excel_button = QPushButton("选择 Excel")
        self.choose_excel_button.setObjectName("secondaryButton")
        self.choose_excel_button.clicked.connect(self.choose_excel)

        self.output_edit = QLineEdit()
        self.output_edit.setReadOnly(True)
        self.output_edit.setText(str(self.output_dir))
        self.choose_output_button = QPushButton("选择输出目录")
        self.choose_output_button.setObjectName("secondaryButton")
        self.choose_output_button.clicked.connect(self.choose_output_dir)
        self.open_output_button = QPushButton("打开输出目录")
        self.open_output_button.setObjectName("secondaryButton")
        self.open_output_button.clicked.connect(self.open_output_dir)

        self.frequency_edit = QLineEdit()
        self.frequency_edit.setPlaceholderText("V1 可填写；留空默认 1e9 Hz；V2 自动读取表格频率")
        self.frequency_unit_combo = QComboBox()
        self.frequency_unit_combo.addItems(FREQUENCY_UNITS)
        self.frequency_unit_combo.setCurrentText("MHz")

        path_layout.addWidget(QLabel("Excel 文件"), 0, 0)
        path_layout.addWidget(self.excel_edit, 0, 1)
        path_layout.addWidget(self.choose_excel_button, 0, 2)
        path_layout.addWidget(QLabel("输出目录"), 1, 0)
        path_layout.addWidget(self.output_edit, 1, 1)
        path_layout.addWidget(self.choose_output_button, 1, 2)
        path_layout.addWidget(self.open_output_button, 1, 3)
        path_layout.addWidget(QLabel("频率"), 2, 0)
        path_layout.addWidget(self.frequency_edit, 2, 1, 1, 2)
        path_layout.addWidget(self.frequency_unit_combo, 2, 3)
        path_layout.setColumnStretch(1, 1)

        sheet_group = QGroupBox("Sheet 选择")
        sheet_layout = QVBoxLayout(sheet_group)
        sheet_layout.setContentsMargins(16, 20, 16, 16)
        sheet_layout.setSpacing(10)
        sheet_buttons = QHBoxLayout()
        select_all_button = QPushButton("全选")
        clear_all_button = QPushButton("取消全选")
        select_all_button.setObjectName("secondaryButton")
        clear_all_button.setObjectName("secondaryButton")
        select_all_button.clicked.connect(lambda: self.set_all_sheets_checked(True))
        clear_all_button.clicked.connect(lambda: self.set_all_sheets_checked(False))
        sheet_buttons.addStretch(1)
        sheet_buttons.addWidget(select_all_button)
        sheet_buttons.addWidget(clear_all_button)

        self.sheet_list = QListWidget()
        self.sheet_list.setAlternatingRowColors(True)
        self.sheet_list.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.sheet_list.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.sheet_list.setUniformItemSizes(True)
        self.sheet_list.itemChanged.connect(self._update_sheet_item_label)
        self.delta_report_checkbox = QCheckBox("生成场景差值图（以第一个选中 Sheet 为基准）")
        self.delta_report_checkbox.setChecked(False)
        self.delta_report_checkbox.stateChanged.connect(self._update_delta_base_label)
        self.delta_base_label = QLabel("当前差值基准：未启用")
        sheet_layout.addLayout(sheet_buttons)
        sheet_layout.addWidget(self.sheet_list)
        sheet_layout.addWidget(self.delta_report_checkbox)
        sheet_layout.addWidget(self.delta_base_label)

        log_group = QGroupBox("转换日志")
        log_layout = QVBoxLayout(log_group)
        log_layout.setContentsMargins(16, 20, 16, 16)
        log_layout.setSpacing(10)
        self.log_output = QTextEdit()
        self.log_output.setReadOnly(True)
        log_layout.addWidget(self.log_output)

        self.convert_button = QPushButton("开始转换")
        self.convert_button.setObjectName("primaryButton")
        self.convert_button.clicked.connect(self.start_conversion)
        self.radar_button = QPushButton("生成雷达图报表")
        self.radar_button.setObjectName("primaryButton")
        self.radar_button.clicked.connect(self.start_radar_report)

        action_row = QHBoxLayout()
        action_row.setSpacing(12)
        action_row.addStretch(1)
        action_row.addWidget(self.convert_button)
        action_row.addWidget(self.radar_button)

        root.addWidget(path_group)
        root.addWidget(sheet_group, 2)
        root.addLayout(action_row)
        root.addWidget(log_group, 3)

        self.setCentralWidget(central)

    def _apply_styles(self) -> None:
        self.setStyleSheet(
            """
            QWidget#appRoot {
                background: #F6F8FB;
                color: #1F2937;
                font-size: 13px;
            }
            QGroupBox {
                background: #FFFFFF;
                border: 1px solid #DCE3EC;
                border-radius: 8px;
                margin-top: 10px;
                font-weight: 600;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 12px;
                padding: 0 6px;
                color: #334155;
            }
            QLabel {
                color: #475569;
            }
            QLineEdit, QTextEdit, QListWidget, QComboBox {
                background: #FFFFFF;
                border: 1px solid #CBD5E1;
                border-radius: 6px;
                padding: 7px 9px;
                selection-background-color: #BFDBFE;
            }
            QLineEdit:focus, QTextEdit:focus, QListWidget:focus, QComboBox:focus {
                border-color: #2563EB;
            }
            QListWidget {
                alternate-background-color: #F8FAFC;
            }
            QListWidget::item {
                min-height: 28px;
                padding: 4px 6px;
            }
            QListView::indicator {
                width: 18px;
                height: 18px;
                border: 1px solid #94A3B8;
                border-radius: 4px;
                background: #FFFFFF;
            }
            QListView::indicator:checked {
                background: #2563EB;
                border-color: #1D4ED8;
            }
            QPushButton {
                border-radius: 6px;
                padding: 8px 14px;
                font-weight: 600;
            }
            QPushButton#primaryButton {
                background: #2563EB;
                border: 1px solid #1D4ED8;
                color: #FFFFFF;
                min-width: 130px;
            }
            QPushButton#primaryButton:hover {
                background: #1D4ED8;
            }
            QPushButton#primaryButton:disabled {
                background: #93C5FD;
                border-color: #93C5FD;
            }
            QPushButton#secondaryButton {
                background: #F8FAFC;
                border: 1px solid #CBD5E1;
                color: #334155;
            }
            QPushButton#secondaryButton:hover {
                background: #EAF2FF;
                border-color: #93C5FD;
            }
            QTextEdit {
                font-family: Consolas, "Courier New", monospace;
                color: #0F172A;
            }
            """
        )

    def choose_excel(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "选择 Excel 文件",
            "",
            "Excel Files (*.xlsx *.xlsm);;All Files (*)",
        )
        if not path:
            return
        self.excel_path = Path(path)
        self.excel_edit.setText(str(self.excel_path))
        self.settings.setValue(LAST_EXCEL_KEY, str(self.excel_path))
        self.last_output_dir = self.output_dir
        self.load_sheet_names()

    def choose_output_dir(self) -> None:
        path = QFileDialog.getExistingDirectory(self, "选择输出目录")
        if not path:
            return
        self.output_dir = Path(path)
        self.last_output_dir = self.output_dir
        self.output_edit.setText(str(self.output_dir))

    def open_output_dir(self) -> None:
        target_dir = self.last_output_dir or self.output_dir
        if target_dir is None:
            return
        target_dir.mkdir(parents=True, exist_ok=True)
        opened = QDesktopServices.openUrl(QUrl.fromLocalFile(str(target_dir)))
        if not opened:
            QMessageBox.warning(self, "打开失败", "无法打开输出目录。")

    def load_sheet_names(self) -> None:
        self.sheet_list.clear()
        self._sheet_selection_order = []
        if self.excel_path is None:
            self._update_delta_base_label()
            return
        try:
            workbook = load_workbook(self.excel_path, read_only=True, data_only=True)
            sheet_names = list(workbook.sheetnames)
            for sheet_name in workbook.sheetnames:
                item = QListWidgetItem(self._sheet_item_text(sheet_name, True))
                item.setData(Qt.UserRole, sheet_name)
                item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
                item.setCheckState(Qt.Checked)
                self.sheet_list.addItem(item)
            self._sheet_selection_order = sheet_names
            workbook.close()
            self._update_sheet_list_height()
            self._update_delta_base_label()
            self.append_log(f"已读取 {self.sheet_list.count()} 个 sheet")
        except Exception as exc:
            self._update_sheet_list_height()
            self._update_delta_base_label()
            QMessageBox.critical(self, "读取失败", str(exc))
            self.append_log(f"[失败] 读取 sheet 列表: {exc}")

    def set_all_sheets_checked(self, checked: bool) -> None:
        state = Qt.Checked if checked else Qt.Unchecked
        for index in range(self.sheet_list.count()):
            self.sheet_list.item(index).setCheckState(state)
        self._sheet_selection_order = (
            [self._sheet_name_for_item(self.sheet_list.item(index)) for index in range(self.sheet_list.count())]
            if checked
            else []
        )
        self._update_delta_base_label()

    def selected_sheets(self) -> list[str]:
        checked_sheets: dict[str, None] = {}
        for index in range(self.sheet_list.count()):
            item = self.sheet_list.item(index)
            if item.checkState() == Qt.Checked:
                checked_sheets[self._sheet_name_for_item(item)] = None
        ordered_sheets = [sheet_name for sheet_name in self._sheet_selection_order if sheet_name in checked_sheets]
        ordered_sheets.extend(sheet_name for sheet_name in checked_sheets if sheet_name not in ordered_sheets)
        return ordered_sheets

    def _update_delta_base_label(self, *_args) -> None:
        if not hasattr(self, "delta_base_label"):
            return
        if not self.delta_report_checkbox.isChecked():
            self.delta_base_label.setText("当前差值基准：未启用")
            return
        sheets = self.selected_sheets()
        if len(sheets) < 2:
            self.delta_base_label.setText("当前差值基准：至少选择 2 个 Sheet")
            return
        self.delta_base_label.setText(f"当前差值基准：{sheets[0]}")

    def start_conversion(self) -> None:
        if self.excel_path is None:
            QMessageBox.warning(self, "缺少 Excel 文件", "请先选择 Excel 文件。")
            return
        if self.output_dir is None:
            QMessageBox.warning(self, "缺少输出目录", "请先选择输出目录。")
            return

        sheets = self.selected_sheets()
        if not sheets:
            QMessageBox.warning(self, "缺少 Sheet", "请至少选择一个 sheet。")
            return

        excel_path = self.excel_path
        output_dir = self.output_dir
        frequency_value = self.frequency_edit.text().strip()
        frequency_unit = self.frequency_unit_combo.currentText()

        def task() -> object:
            return convert_excel(
                excel_path,
                output_dir,
                sheets,
                frequency_value=frequency_value,
                frequency_unit=frequency_unit,
            )

        self._start_background_task(
            "开始转换...",
            "转换失败",
            task,
            self._handle_conversion_result,
        )

    def start_radar_report(self) -> None:
        if self.excel_path is None:
            QMessageBox.warning(self, "缺少 Excel 文件", "请先选择 Excel 文件。")
            return
        if self.output_dir is None:
            QMessageBox.warning(self, "缺少输出目录", "请先选择输出目录。")
            return

        sheets = self.selected_sheets()
        if not sheets:
            QMessageBox.warning(self, "缺少 Sheet", "请至少选择一个 sheet。")
            return

        excel_path = self.excel_path
        output_dir = self.output_dir
        include_delta = self.delta_report_checkbox.isChecked()

        def task() -> object:
            return generate_radar_report(
                excel_path,
                output_dir,
                sheets,
                include_delta=include_delta,
            )

        self._start_background_task(
            "开始生成雷达图报表...",
            "报表生成失败",
            task,
            self._handle_radar_report_result,
        )

    def append_log(self, message: str) -> None:
        self.log_output.append(message)

    def closeEvent(self, event) -> None:
        if self._task_thread is not None:
            QMessageBox.warning(self, "任务运行中", "当前任务还未完成，请等待处理结束后再关闭窗口。")
            event.ignore()
            return
        super().closeEvent(event)

    def _start_background_task(
        self,
        start_message: str,
        error_title: str,
        task: Callable[[], object],
        on_success: Callable[[object], None],
    ) -> None:
        if self._task_thread is not None:
            QMessageBox.warning(self, "任务运行中", "当前任务还未完成，请稍后再试。")
            return

        self._set_task_controls_enabled(False)
        QApplication.setOverrideCursor(Qt.WaitCursor)
        self.append_log(start_message)
        self.append_log("正在后台处理，界面会保持可响应...")

        thread = QThread(self)
        worker = _TaskWorker(task)
        worker.moveToThread(thread)

        self._task_thread = thread
        self._task_worker = worker
        self._task_success_handler = on_success
        self._task_error_title = error_title

        thread.started.connect(worker.run)
        worker.succeeded.connect(self._handle_background_success)
        worker.failed.connect(self._handle_background_failure)
        worker.done.connect(self._handle_background_done)
        worker.done.connect(thread.quit)
        worker.done.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        thread.finished.connect(self._clear_background_task)
        thread.start()

    @Slot(object)
    def _handle_background_success(self, result: object) -> None:
        if self._task_success_handler is None:
            return
        try:
            self._task_success_handler(result)
        except Exception:
            details = traceback.format_exc()
            self.append_log(f"[失败] UI 更新: {details}")
            QMessageBox.critical(self, self._task_error_title or "任务失败", details)

    @Slot(str)
    def _handle_background_failure(self, details: str) -> None:
        self.append_log(f"[失败] {details}")
        QMessageBox.critical(self, self._task_error_title or "任务失败", details)

    @Slot()
    def _handle_background_done(self) -> None:
        QApplication.restoreOverrideCursor()
        self._set_task_controls_enabled(True)
        self.append_log("任务处理结束。")
        self._task_success_handler = None
        self._task_error_title = ""

    @Slot()
    def _clear_background_task(self) -> None:
        self._task_thread = None
        self._task_worker = None

    def _set_task_controls_enabled(self, enabled: bool) -> None:
        self.choose_excel_button.setEnabled(enabled)
        self.choose_output_button.setEnabled(enabled)
        self.frequency_edit.setEnabled(enabled)
        self.frequency_unit_combo.setEnabled(enabled)
        self.sheet_list.setEnabled(enabled)
        self.delta_report_checkbox.setEnabled(enabled)
        self.convert_button.setEnabled(enabled)
        self.radar_button.setEnabled(enabled)

    def _handle_conversion_result(self, result: object) -> None:
        for line in result.log_lines:
            self.append_log(line)
        if result.log_path is not None:
            self.append_log(f"日志文件: {result.log_path}")
        self.last_output_dir = self._conversion_output_dir(result.generated_files)
        self.append_log(f"可打开目录: {self.last_output_dir}")
        QMessageBox.information(self, "转换完成", f"转换完成，生成 {result.generated_count} 个文件。")

    def _handle_radar_report_result(self, result: object) -> None:
        self.append_log(f"解析到矩阵数量: {result.matrix_count}")
        self.append_log(f"生成单图数量: {result.single_chart_count}")
        self.append_log(f"生成对比图数量: {result.compare_chart_count}")
        self.append_log(f"生成差值图数量: {result.delta_chart_count}")
        self.append_log(f"雷达图报表: {result.output_path}")
        self.last_output_dir = result.output_path.parent
        self.append_log(f"可打开目录: {self.last_output_dir}")
        QMessageBox.information(self, "报表生成完成", f"雷达图报表已生成:\n{result.output_path}")

    def _update_sheet_item_label(self, item: QListWidgetItem) -> None:
        if self._updating_sheet_item:
            return
        self._updating_sheet_item = True
        try:
            sheet_name = self._sheet_name_for_item(item)
            checked = item.checkState() == Qt.Checked
            if checked and sheet_name not in self._sheet_selection_order:
                self._sheet_selection_order.append(sheet_name)
            elif not checked:
                self._sheet_selection_order = [name for name in self._sheet_selection_order if name != sheet_name]
            item.setText(self._sheet_item_text(sheet_name, checked))
            item.setData(Qt.UserRole, sheet_name)
            self._update_delta_base_label()
        finally:
            self._updating_sheet_item = False

    def _sheet_name_for_item(self, item: QListWidgetItem) -> str:
        sheet_name = item.data(Qt.UserRole)
        if sheet_name:
            return str(sheet_name)
        return item.text().removeprefix(SHEET_CHECKED_PREFIX).removeprefix(SHEET_UNCHECKED_PREFIX)

    def _sheet_item_text(self, sheet_name: str, checked: bool) -> str:
        prefix = SHEET_CHECKED_PREFIX if checked else SHEET_UNCHECKED_PREFIX
        return f"{prefix}{sheet_name}"

    def _update_sheet_list_height(self) -> None:
        count = self.sheet_list.count()
        if count == 0:
            visible_rows = MIN_VISIBLE_SHEET_ROWS
            scroll_policy = Qt.ScrollBarAlwaysOff
        else:
            visible_rows = min(max(count, MIN_VISIBLE_SHEET_ROWS), MAX_VISIBLE_SHEET_ROWS)
            scroll_policy = Qt.ScrollBarAsNeeded if count > MAX_VISIBLE_SHEET_ROWS else Qt.ScrollBarAlwaysOff

        row_height = self.sheet_list.sizeHintForRow(0) if count else 30
        if row_height <= 0:
            row_height = 30
        height = row_height * visible_rows + self.sheet_list.frameWidth() * 2 + 10
        self.sheet_list.setFixedHeight(height)
        self.sheet_list.setVerticalScrollBarPolicy(scroll_policy)

    def _load_last_excel(self) -> None:
        saved_path = self.settings.value(LAST_EXCEL_KEY, "", str)
        if not saved_path:
            return

        excel_path = Path(saved_path)
        if not excel_path.is_file():
            self.settings.remove(LAST_EXCEL_KEY)
            self.excel_path = None
            self.excel_edit.clear()
            return

        self.excel_path = excel_path
        self.excel_edit.setText(str(excel_path))
        self.load_sheet_names()

    def _conversion_output_dir(self, generated_files: list[Path]) -> Path:
        if generated_files:
            return generated_files[0].parent
        if self.output_dir is None:
            return get_default_output_dir()
        if self.excel_path is None:
            return self.output_dir
        return self.output_dir / sanitize_filename(self.excel_path.stem)


def main() -> None:
    app = QApplication([])
    window = MainWindow()
    window.show()
    app.exec()
