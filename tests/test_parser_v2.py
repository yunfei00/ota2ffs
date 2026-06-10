from openpyxl import Workbook

from ota2ffs import parser_v2


def _add_table(ws, start_row, polarization, offset):
    ws.cell(row=start_row, column=1, value="Polarization")
    ws.cell(row=start_row, column=2, value=polarization)
    ws.cell(row=start_row, column=6, value="900 MHz")
    ws.cell(row=start_row + 1, column=1, value="Phi\\Theta")
    for column, theta in enumerate([0, 90, 180], start=2):
        ws.cell(row=start_row + 1, column=column, value=theta)
    for row_offset, phi in enumerate([0, 180], start=2):
        row = start_row + row_offset
        ws.cell(row=row, column=1, value=phi)
        for column_offset, _theta in enumerate([0, 90, 180], start=2):
            ws.cell(row=row, column=column_offset, value=offset + row_offset + column_offset)


def _build_v2_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "V2Sheet"
    _add_table(ws, 1, "Theta", 10)
    _add_table(ws, 8, "Phi", 20)
    _add_table(ws, 15, "Total", 30)
    return ws


def test_v2_parser_builds_regular_and_total_sources():
    ws = _build_v2_sheet()

    assert parser_v2.is_v2_sheet(ws)
    sources = parser_v2.parse_sheet(ws)

    assert len(sources) == 2
    regular, total = sources
    assert regular.version == "V2"
    assert regular.frequency_mhz == 900
    assert regular.theta_angles == [0, 90, 180]
    assert regular.phi_angles == [0, 180]
    assert regular.e_theta_db[(0, 0)] == 14
    assert regular.e_phi_db[(0, 0)] == 24
    assert total.suffix == "_拓图"
    assert total.e_theta_db[(0, 0)] == 34
    assert total.e_phi_db == {}
