from pathlib import Path

import pytest
from openpyxl import Workbook

from ota2ffs.converter import convert_excel


def _add_v2_table(ws, start_row, polarization, value, start_column=1, frequency=1800):
    ws.cell(row=start_row, column=start_column, value="Polarization")
    ws.cell(row=start_row, column=start_column + 1, value=polarization)
    ws.cell(row=start_row, column=start_column + 2, value=frequency)
    ws.cell(row=start_row + 1, column=start_column, value="Phi\\Theta")
    ws.cell(row=start_row + 1, column=start_column + 1, value=0)
    ws.cell(row=start_row + 1, column=start_column + 2, value=180)
    ws.cell(row=start_row + 2, column=start_column, value=0)
    ws.cell(row=start_row + 2, column=start_column + 1, value=value)
    ws.cell(row=start_row + 2, column=start_column + 2, value=value)


def _add_v2_sheet(workbook, title, base_value):
    ws = workbook.create_sheet(title)
    _add_v2_table(ws, 1, "Theta", base_value)
    _add_v2_table(ws, 6, "Phi", base_value + 1)
    _add_v2_table(ws, 11, "Total", base_value + 2)
    return ws


def _add_shifted_v2_sheet(workbook, title, base_value):
    ws = workbook.create_sheet(title)
    _add_v2_table(ws, 4, "Theta", base_value, start_column=3, frequency=2450)
    _add_v2_table(ws, 12, "Phi", base_value + 1, start_column=6, frequency=2450)
    _add_v2_table(ws, 22, "Total", base_value + 2, start_column=2, frequency=2450)
    return ws


def _add_v1_block(ws, start_row, start_column, title, values_offset):
    theta_angles = [0, 90, 180]
    phi_angles = [-180, 0]

    ws.cell(row=start_row, column=start_column, value=title)
    ws.cell(row=start_row, column=start_column + 1, value="Phi Angle")
    ws.cell(row=start_row + 1, column=start_column + 1, value="Theta Angle")
    for column_offset, theta in enumerate(theta_angles, start=2):
        ws.cell(row=start_row, column=start_column + column_offset, value=theta)
    for row_offset, phi in enumerate(phi_angles, start=2):
        row = start_row + row_offset
        ws.cell(row=row, column=start_column + 1, value=phi)
        for column_offset, _theta in enumerate(theta_angles, start=2):
            ws.cell(row=row, column=start_column + column_offset, value=values_offset + row_offset + column_offset)


def _add_shifted_v1_sheet(workbook, title):
    ws = workbook.create_sheet(title)
    _add_v1_block(ws, 3, 4, "Theta", -30)
    _add_v1_block(ws, 18, 7, "Phi", -40)
    return ws


def test_converter_keeps_going_when_one_sheet_fails(tmp_path):
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Valid"
    _add_v2_table(ws, 1, "Theta", 20)
    _add_v2_table(ws, 6, "Phi", 6)
    _add_v2_table(ws, 11, "Total", 3)

    invalid = workbook.create_sheet("Invalid")
    invalid["A1"] = "not a known format"

    excel_path = tmp_path / "input.xlsx"
    workbook.save(excel_path)

    output_dir = tmp_path / "out"
    log_dir = tmp_path / "log"
    result = convert_excel(
        excel_path,
        output_dir,
        ["Valid", "Invalid"],
        frequency_value="9999",
        frequency_unit="GHz",
        log_dir=log_dir,
    )

    assert result.generated_count == 4
    assert "Invalid" in result.failures
    excel_output_dir = output_dir / "input"
    assert (excel_output_dir / "Valid_Rx.ffs").exists()
    assert (excel_output_dir / "Valid_Tx.ffs").exists()
    assert (excel_output_dir / "Valid_total_Rx.ffs").exists()
    assert (excel_output_dir / "Valid_total_Tx.ffs").exists()
    assert result.log_path is not None
    assert result.log_path.exists()
    assert result.log_path.parent == log_dir

    rx_lines = (excel_output_dir / "Valid_Rx.ffs").read_text(encoding="utf-8").splitlines()
    tx_lines = (excel_output_dir / "Valid_Tx.ffs").read_text(encoding="utf-8").splitlines()
    assert rx_lines[0] == "// CST Farfield Source File"
    assert tx_lines[0] == "// CST Farfield Source File"
    assert rx_lines[9] == "1"
    assert tx_lines[9] == "1"
    assert rx_lines[24] == "1800000000"
    assert tx_lines[24] == "1800000000"
    assert rx_lines[26] == "// >> Total #phi samples, total #theta samples"
    assert tx_lines[26] == "// >> Total #phi samples, total #theta samples"
    assert rx_lines[27] == "1   2"
    assert tx_lines[27] == "1   2"
    assert rx_lines[29] == "// >> Phi, Theta, Re(E_Theta), Im(E_Theta), Re(E_Phi), Im(E_Phi):"
    assert tx_lines[29] == "// >> Phi, Theta, Re(E_Theta), Im(E_Theta), Re(E_Phi), Im(E_Phi):"

    rx_line = rx_lines[30]
    tx_line = tx_lines[30]

    assert "," not in rx_line
    assert "," not in tx_line
    assert float(rx_line.split()[2]) == pytest.approx(10)
    assert float(tx_line.split()[2]) == pytest.approx(0.1)


def test_converter_processes_multiple_selected_sheets(tmp_path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_v2_sheet(workbook, "SheetA", 20)
    _add_v2_sheet(workbook, "SheetB", 40)
    invalid = workbook.create_sheet("Broken")
    invalid["A1"] = "not a known format"

    excel_path = tmp_path / "multi.xlsx"
    workbook.save(excel_path)

    output_dir = tmp_path / "out"
    result = convert_excel(
        excel_path,
        output_dir,
        ["SheetA", "SheetB", "Broken"],
        log_dir=tmp_path / "log",
    )

    excel_output_dir = output_dir / "multi"
    assert result.generated_count == 8
    assert result.failures == {"Broken": "无法识别为 V1 或 V2 格式"}

    expected_files = {
        "SheetA_Rx.ffs",
        "SheetA_Tx.ffs",
        "SheetA_total_Rx.ffs",
        "SheetA_total_Tx.ffs",
        "SheetB_Rx.ffs",
        "SheetB_Tx.ffs",
        "SheetB_total_Rx.ffs",
        "SheetB_total_Tx.ffs",
    }
    assert {path.name for path in result.generated_files} == expected_files
    for filename in expected_files:
        assert (excel_output_dir / filename).exists()


def test_converter_processes_mixed_shifted_v1_and_v2_sheets_with_default_frequency(tmp_path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_shifted_v1_sheet(workbook, "OffsetV1")
    _add_shifted_v2_sheet(workbook, "OffsetV2", 20)

    excel_path = tmp_path / "mixed.xlsx"
    workbook.save(excel_path)

    output_dir = tmp_path / "out"
    result = convert_excel(excel_path, output_dir, log_dir=tmp_path / "log")

    excel_output_dir = output_dir / "mixed"
    assert result.failures == {}
    assert result.generated_count == 6
    assert (excel_output_dir / "OffsetV1_Rx.ffs").exists()
    assert (excel_output_dir / "OffsetV1_Tx.ffs").exists()
    assert (excel_output_dir / "OffsetV2_Rx.ffs").exists()
    assert (excel_output_dir / "OffsetV2_Tx.ffs").exists()
    assert (excel_output_dir / "OffsetV2_total_Rx.ffs").exists()
    assert (excel_output_dir / "OffsetV2_total_Tx.ffs").exists()

    v1_lines = (excel_output_dir / "OffsetV1_Rx.ffs").read_text(encoding="utf-8").splitlines()
    v2_lines = (excel_output_dir / "OffsetV2_Rx.ffs").read_text(encoding="utf-8").splitlines()
    assert v1_lines[24] == "1000000000"
    assert v2_lines[24] == "2450000000"


def test_sample_mixed_multisheet_excel_converts(tmp_path):
    sample_path = Path(__file__).resolve().parents[1] / "samples" / "ota_mixed_multisheet_sample.xlsx"

    result = convert_excel(sample_path, tmp_path / "out", log_dir=tmp_path / "log")

    assert result.failures == {}
    assert result.generated_count == 6


def test_sample_v1_default_frequency_excel_converts(tmp_path):
    sample_path = Path(__file__).resolve().parents[1] / "samples" / "ota_v1_default_frequency_sample.xlsx"

    result = convert_excel(sample_path, tmp_path / "out", log_dir=tmp_path / "log")

    assert result.failures == {}
    assert result.generated_count == 2
    output_file = tmp_path / "out" / "ota_v1_default_frequency_sample" / "V1_Default_1e9_Rx.ffs"
    assert output_file.read_text(encoding="utf-8").splitlines()[24] == "1000000000"
