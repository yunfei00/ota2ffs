import hashlib
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from ota2ffs.matrix_model import PatternMatrix
from ota2ffs.radar_report import add_col_charts, add_row_charts, generate_radar_report


def _matrix(sheet_name="S1", block_name="Theta"):
    return PatternMatrix(
        sheet_name=sheet_name,
        block_name=block_name,
        row_label="Phi Angle",
        col_label="Theta Angle",
        row_angles=[0, 180],
        col_angles=[0, 90, 180],
        values=[[-10, -20, -30], [-40, -50, -60]],
    )


def _add_v2_table(ws, start_row, start_column, block_name, base_value):
    ws.cell(row=start_row, column=start_column, value="Polarization")
    ws.cell(row=start_row, column=start_column + 1, value=block_name)
    ws.cell(row=start_row, column=start_column + 2, value=1800)
    ws.cell(row=start_row + 1, column=start_column, value="Phi\\Theta")
    ws.cell(row=start_row + 1, column=start_column + 1, value=0)
    ws.cell(row=start_row + 1, column=start_column + 2, value=180)
    ws.cell(row=start_row + 2, column=start_column, value=0)
    ws.cell(row=start_row + 2, column=start_column + 1, value=base_value)
    ws.cell(row=start_row + 2, column=start_column + 2, value=base_value - 1)
    ws.cell(row=start_row + 3, column=start_column, value=180)
    ws.cell(row=start_row + 3, column=start_column + 1, value=base_value - 2)
    ws.cell(row=start_row + 3, column=start_column + 2, value=base_value - 3)


def _add_v2_sheet(workbook, title, base_value):
    ws = workbook.create_sheet(title)
    _add_v2_table(ws, 1, 1, "Theta", base_value)
    _add_v2_table(ws, 8, 4, "Phi", base_value - 10)
    _add_v2_table(ws, 16, 2, "Total", base_value - 20)


def _add_v1_block(ws, start_row, start_column, block_name, base_value):
    ws.cell(row=start_row, column=start_column, value=block_name)
    ws.cell(row=start_row, column=start_column + 1, value="Phi Angle")
    ws.cell(row=start_row + 1, column=start_column + 1, value="Theta Angle")
    for column_offset, theta in enumerate([0, 180], start=2):
        ws.cell(row=start_row, column=start_column + column_offset, value=theta)
    for row_offset, phi in enumerate([-180, 0], start=2):
        row = start_row + row_offset
        ws.cell(row=row, column=start_column + 1, value=phi)
        for column_offset, _theta in enumerate([0, 180], start=2):
            ws.cell(row=row, column=start_column + column_offset, value=base_value - row_offset - column_offset)


def _add_v1_sheet(workbook, title, base_value):
    ws = workbook.create_sheet(title)
    _add_v1_block(ws, 1, 1, "Theta", base_value)
    _add_v1_block(ws, 8, 1, "Phi", base_value - 10)


def test_row_chart_count_equals_row_angle_count():
    workbook = Workbook()
    report_ws = workbook.active
    data_ws = workbook.create_sheet("Data")
    counts = {"single": 0, "compare": 0}

    add_row_charts(report_ws, data_ws, _matrix(), 1, 1, {"row": 1}, counts)

    assert counts["single"] == 2
    assert len(report_ws._charts) == 2


def test_col_chart_count_equals_col_angle_count():
    workbook = Workbook()
    report_ws = workbook.active
    data_ws = workbook.create_sheet("Data")
    counts = {"single": 0, "compare": 0}

    add_col_charts(report_ws, data_ws, _matrix(), 1, 1, {"row": 1}, counts)

    assert counts["single"] == 3
    assert len(report_ws._charts) == 3


def test_radar_charts_use_excel_native_defaults(tmp_path):
    workbook = Workbook()
    report_ws = workbook.active
    data_ws = workbook.create_sheet("Data")
    counts = {"single": 0, "compare": 0}

    add_row_charts(report_ws, data_ws, _matrix(), 1, 1, {"row": 1}, counts)

    chart = report_ws._charts[0]
    assert chart.style == 26
    assert chart.legend is not None
    assert chart.legend.legendPos == "r"
    assert chart.dataLabels is None

    output_path = tmp_path / "chart_native.xlsx"
    workbook.save(output_path)
    with zipfile.ZipFile(output_path) as archive:
        chart_xml = archive.read("xl/charts/chart1.xml").decode("utf-8")

    assert '<style val="26"' in chart_xml
    assert '<radarStyle val="standard"' in chart_xml
    assert "<legend>" in chart_xml
    assert "<dLbls>" not in chart_xml
    assert "<tickLblPos" not in chart_xml


def test_multiple_sheets_generate_compare_charts_and_output_xlsx(tmp_path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_v2_sheet(workbook, "S1", -10)
    _add_v2_sheet(workbook, "S2", -20)
    excel_path = tmp_path / "radar_input.xlsx"
    workbook.save(excel_path)

    result = generate_radar_report(excel_path, tmp_path / "out", ["S1", "S2"])

    assert result.output_path.exists()
    assert result.output_path.name == "radar_input_Radar_Report.xlsx"
    assert result.matrix_count == 6
    assert result.single_chart_count == 0
    assert result.compare_chart_count == 12


def test_compare_report_layout_keeps_blocks_vertical_and_compare_data_right(tmp_path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_v2_sheet(workbook, "S1", -10)
    _add_v2_sheet(workbook, "S2", -20)
    excel_path = tmp_path / "layout_input.xlsx"
    workbook.save(excel_path)

    result = generate_radar_report(excel_path, tmp_path / "out", ["S1", "S2"])

    output_workbook = load_workbook(result.output_path, data_only=False)
    try:
        data_positions = _positions_by_value(output_workbook["Normalized_Data"])
        report_positions = _positions_by_value(output_workbook["Radar_Report"])

        data_block_rows = [data_positions[f"Block: {block}"][0] for block in ("Theta", "Phi", "Total")]
        report_block_rows = [report_positions[f"Block: {block}"][0] for block in ("Theta", "Phi", "Total")]
        assert data_block_rows == sorted(data_block_rows)
        assert report_block_rows == sorted(report_block_rows)

        for block in ("Theta", "Phi", "Total"):
            source_row, source_col = data_positions[f"Block: {block}"]
            compare_row, compare_col = data_positions[f"Compare Data: {block}"]
            assert compare_row == source_row
            assert compare_col > source_col
    finally:
        output_workbook.close()


def test_compare_data_cells_link_to_source_matrix_cells(tmp_path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_v2_sheet(workbook, "S1", -10)
    _add_v2_sheet(workbook, "S2", -20)
    excel_path = tmp_path / "linked_input.xlsx"
    workbook.save(excel_path)

    result = generate_radar_report(excel_path, tmp_path / "out", ["S1", "S2"])

    output_workbook = load_workbook(result.output_path, data_only=False)
    try:
        ws = output_workbook["Normalized_Data"]
        positions = _positions_by_value(ws)
        source_row, source_col = positions["Sheet: S1 / Block: Theta"]
        compare_row, compare_col = positions["Compare_Theta_Row_0"]
        expected_source_cell = f"${get_column_letter(source_col + 1)}${source_row + 2}"
        compare_value_cell = ws.cell(row=compare_row + 2, column=compare_col + 1)

        assert compare_value_cell.data_type == "f"
        assert compare_value_cell.value == f"={expected_source_cell}"
        assert output_workbook.calculation.calcMode == "auto"
        assert output_workbook.calculation.forceFullCalc
    finally:
        output_workbook.close()


def test_mixed_v1_v2_compare_uses_v1_output_phi_angles(tmp_path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_v2_sheet(workbook, "V2", -10)
    _add_v1_sheet(workbook, "V1", -20)
    excel_path = tmp_path / "mixed_compare.xlsx"
    workbook.save(excel_path)

    result = generate_radar_report(excel_path, tmp_path / "out", ["V2", "V1"])

    output_workbook = load_workbook(result.output_path, data_only=False)
    try:
        positions = _positions_by_value(output_workbook["Normalized_Data"])

        assert "Compare_Theta_Row_0" in positions
        assert "Compare_Theta_Row_180" in positions
        assert "Compare_Theta_Row_-180" not in positions
        assert result.matrix_count == 5
        assert result.compare_chart_count == 8
    finally:
        output_workbook.close()


def test_radar_report_does_not_modify_source_excel(tmp_path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_v2_sheet(workbook, "S1", -10)
    excel_path = tmp_path / "source.xlsx"
    workbook.save(excel_path)
    before_hash = _sha256(excel_path)

    generate_radar_report(excel_path, tmp_path / "out", ["S1"])

    assert _sha256(excel_path) == before_hash


def test_v2_sample_sheets_are_merged_into_compare_charts(tmp_path):
    sample_path = Path(__file__).resolve().parents[1] / "samples" / "ota_v2_sample.xlsx"

    result = generate_radar_report(sample_path, tmp_path / "out", ["V2_Sample", "V2_Sample_Compare"])

    workbook = load_workbook(result.output_path, data_only=False)
    try:
        assert result.matrix_count == 6
        assert result.single_chart_count == 0
        assert result.compare_chart_count == 42
        assert len(workbook["Radar_Report"]._charts) == 42
    finally:
        workbook.close()


def _sha256(path):
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _positions_by_value(ws):
    positions = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                positions[cell.value] = (cell.row, cell.column)
    return positions
