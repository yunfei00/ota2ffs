from openpyxl import Workbook

from ota2ffs import parser_v1


def _build_v1_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "V1Sheet"

    theta_angles = [0, 90, 180]
    phi_angles = [-180, 0]

    ws["A1"] = "Theta"
    ws["B1"] = "Phi Angle"
    ws["B2"] = "Theta Angle"
    for index, theta in enumerate(theta_angles, start=3):
        ws.cell(row=1, column=index, value=theta)
    for row_index, phi in enumerate(phi_angles, start=3):
        ws.cell(row=row_index, column=2, value=phi)
        for col_index, theta in enumerate(theta_angles, start=3):
            ws.cell(row=row_index, column=col_index, value=-(row_index + col_index))

    ws["A30"] = "Phi"
    ws["B30"] = "Phi Angle"
    ws["B31"] = "Theta Angle"
    for index, theta in enumerate(theta_angles, start=3):
        ws.cell(row=30, column=index, value=theta)
    for row_index, phi in enumerate(phi_angles, start=32):
        ws.cell(row=row_index, column=2, value=phi)
        for col_index, theta in enumerate(theta_angles, start=3):
            ws.cell(row=row_index, column=col_index, value=-(row_index + col_index + 10))

    return ws


def _add_v1_block(ws, start_row, start_column, title, values_offset):
    theta_angles = [0, 90, 180]
    phi_angles = [-180, 0]

    ws.cell(row=start_row, column=start_column, value=title)
    ws.cell(row=start_row, column=start_column + 1, value="Phi Angle")
    ws.cell(row=start_row + 1, column=start_column + 1, value="Theta Angle")
    for index, theta in enumerate(theta_angles, start=start_column + 2):
        ws.cell(row=start_row, column=index, value=theta)
    for row_offset, phi in enumerate(phi_angles, start=2):
        row = start_row + row_offset
        ws.cell(row=row, column=start_column + 1, value=phi)
        for column_offset, _theta in enumerate(theta_angles, start=2):
            ws.cell(row=row, column=start_column + column_offset, value=values_offset + row_offset + column_offset)


def test_v1_parser_detects_and_completes_angles():
    ws = _build_v1_sheet()

    assert parser_v1.is_v1_sheet(ws)
    source = parser_v1.parse_sheet(ws)

    assert source.version == "V1"
    assert source.theta_angles == [0, 90, 180]
    assert source.phi_angles == [0, 180, 360]
    assert source.e_theta_db[(180, 90)] == -8
    assert (360, 90) not in source.e_theta_db


def test_v1_parser_supports_shifted_blocks():
    ws = _build_v1_sheet()
    ws.delete_rows(1, ws.max_row)

    _add_v1_block(ws, 5, 4, "Theta", 10)
    _add_v1_block(ws, 18, 7, "Phi", 20)

    assert parser_v1.is_v1_sheet(ws)
    source = parser_v1.parse_sheet(ws)

    assert source.theta_angles == [0, 90, 180]
    assert source.phi_angles == [0, 180, 360]
    assert source.e_theta_db[(0, 0)] == 14
    assert source.e_phi_db[(180, 90)] == 26


def test_v1_parser_supports_total_block_and_loose_angle_labels():
    ws = _build_v1_sheet()
    ws["B1"] = "Phi Angle (deg)"
    ws["B2"] = "Theta Angle °"
    _add_v1_block(ws, 45, 4, "Total", -100)

    assert parser_v1.is_v1_sheet(ws)
    sources = parser_v1.parse_sources(ws)

    assert len(sources) == 2
    regular, total = sources
    assert regular.e_theta_db[(180, 90)] == -8
    assert total.suffix == "_total"
    assert total.e_theta_db[(0, 0)] == -96
    assert total.e_phi_db == {}
