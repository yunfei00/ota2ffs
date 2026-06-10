from openpyxl import Workbook

from ota2ffs.matrix_parser import parse_v1_sheet, parse_v2_sheet


def _add_v1_block(ws, start_row, start_column, block_name, base_value):
    ws.cell(row=start_row, column=start_column, value=block_name)
    ws.cell(row=start_row, column=start_column + 1, value="Phi Angle")
    ws.cell(row=start_row + 1, column=start_column + 1, value="Theta Angle")
    for column_offset, theta in enumerate([0, 90], start=2):
        ws.cell(row=start_row, column=start_column + column_offset, value=theta)
    for row_offset, phi in enumerate([-180, 0], start=2):
        row = start_row + row_offset
        ws.cell(row=row, column=start_column + 1, value=phi)
        for column_offset, _theta in enumerate([0, 90], start=2):
            ws.cell(row=row, column=start_column + column_offset, value=base_value - row_offset - column_offset)


def _add_v2_table(ws, start_row, start_column, block_name, base_value):
    ws.cell(row=start_row, column=start_column, value="Polarization")
    ws.cell(row=start_row, column=start_column + 1, value=block_name)
    ws.cell(row=start_row + 1, column=start_column, value="Phi\\Theta")
    for column_offset, theta in enumerate([0, 90], start=1):
        ws.cell(row=start_row + 1, column=start_column + column_offset, value=theta)
    for row_offset, phi in enumerate([0, 180], start=2):
        row = start_row + row_offset
        ws.cell(row=row, column=start_column, value=phi)
        for column_offset, _theta in enumerate([0, 90], start=1):
            ws.cell(row=row, column=start_column + column_offset, value=base_value - row_offset - column_offset)


def test_v1_sheet_parses_theta_and_phi_matrices():
    workbook = Workbook()
    ws = workbook.active
    ws.title = "V1"
    _add_v1_block(ws, 3, 4, "Theta", -10)
    _add_v1_block(ws, 20, 7, "Phi", -20)

    matrices = parse_v1_sheet(ws)

    assert [matrix.block_name for matrix in matrices] == ["Theta", "Phi"]
    assert matrices[0].sheet_name == "V1"
    assert matrices[0].row_angles == [-180, 0]
    assert matrices[0].col_angles == [0, 90]


def test_v2_sheet_parses_theta_phi_total_matrices():
    workbook = Workbook()
    ws = workbook.active
    ws.title = "V2"
    _add_v2_table(ws, 2, 3, "Theta", -10)
    _add_v2_table(ws, 10, 5, "Phi", -20)
    _add_v2_table(ws, 18, 2, "Total", -30)

    matrices = parse_v2_sheet(ws)

    assert [matrix.block_name for matrix in matrices] == ["Theta", "Phi", "Total"]
    assert matrices[0].row_angles == [0, 180]
    assert matrices[0].col_angles == [0, 90]


def test_negative_values_are_normalized_to_positive_values():
    workbook = Workbook()
    ws = workbook.active
    _add_v2_table(ws, 1, 1, "Theta", -10)
    _add_v2_table(ws, 8, 1, "Phi", -20)

    matrix = parse_v2_sheet(ws)[0]

    assert matrix.values[0][0] == -13
    assert matrix.normalized_values()[0][0] == 13
